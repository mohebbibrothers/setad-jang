"""Excel export helpers for Kindness Wall admin analytics.

The exports are generated in-memory with openpyxl, RTL-friendly styling, frozen
headers, deterministic filenames, and no view-layer formatting logic.
"""

from __future__ import annotations

from collections.abc import Iterable
from io import BytesIO

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from apps.kindness_wall.models import KindnessListing, KindnessListingReport

_HEADER_FILL = PatternFill("solid", fgColor="7A277A")
_HEADER_FONT = Font(name="Tahoma", bold=True, color="FFFFFF")
_BODY_FONT = Font(name="Tahoma", size=11)
_SUMMARY_FILL = PatternFill("solid", fgColor="F2EAF2")


def build_listings_workbook(*, listings: Iterable[KindnessListing]) -> BytesIO:
    """Build an RTL Excel workbook for Kindness Wall listings."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "آگهی‌ها"
    _prepare_sheet(worksheet)
    headers = [
        "شناسه",
        "نوع",
        "وضعیت",
        "عنوان",
        "دسته‌بندی",
        "استان",
        "شهر",
        "نام صاحب آگهی",
        "شماره تماس Snapshot",
        "بازدید",
        "نمایش شماره",
        "ذخیره",
        "گزارش",
        "تاریخ انتشار",
        "تاریخ انقضا",
        "تاریخ ایجاد",
    ]
    worksheet.append(headers)
    _style_header(worksheet)

    total_views = 0
    total_reveals = 0
    for listing in listings:
        total_views += listing.view_count
        total_reveals += listing.contact_reveal_count
        worksheet.append(
            [
                listing.pk,
                listing.get_listing_type_display(),
                listing.get_status_display(),
                listing.title,
                listing.category.title,
                listing.province,
                listing.city,
                listing.owner_full_name_snapshot,
                listing.contact_phone_snapshot,
                listing.view_count,
                listing.contact_reveal_count,
                listing.bookmark_count,
                listing.report_count,
                _format_dt(listing.published_at),
                _format_dt(listing.expires_at),
                _format_dt(listing.created_at),
            ]
        )

    worksheet.append(["جمع", "", "", "", "", "", "", "", "", total_views, total_reveals, "", "", "", "", ""])
    _style_body_and_summary(worksheet)
    _set_widths(worksheet, [12, 20, 18, 34, 24, 16, 16, 24, 24, 14, 14, 12, 12, 22, 22, 22])
    return _save_workbook(workbook)


def build_reports_workbook(*, reports: Iterable[KindnessListingReport]) -> BytesIO:
    """Build an RTL Excel workbook for Kindness Wall report moderation."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "گزارش‌ها"
    _prepare_sheet(worksheet)
    headers = [
        "شناسه گزارش",
        "شناسه آگهی",
        "عنوان آگهی",
        "دلیل",
        "وضعیت",
        "توضیحات کاربر",
        "یادداشت ادمین",
        "گزارش‌دهنده",
        "بررسی‌کننده",
        "تاریخ ثبت",
        "تاریخ بررسی",
    ]
    worksheet.append(headers)
    _style_header(worksheet)

    for report in reports:
        worksheet.append(
            [
                report.pk,
                report.listing_id,
                report.listing.title,
                report.get_reason_display(),
                report.get_status_display(),
                report.description,
                report.admin_note,
                getattr(report.reported_by, "full_name", "") or str(report.reported_by),
                getattr(report.reviewed_by, "full_name", "") if report.reviewed_by else "",
                _format_dt(report.created_at),
                _format_dt(report.reviewed_at),
            ]
        )

    _style_body_and_summary(worksheet, include_summary=False)
    _set_widths(worksheet, [14, 14, 34, 22, 18, 42, 42, 24, 24, 22, 22])
    return _save_workbook(workbook)


def build_kindness_export_filename(*, export_type: str) -> str:
    """Build deterministic Kindness Wall export filenames."""
    date_part = timezone.now().strftime("%Y%m%d")
    return f"kindness-wall-{export_type}-{date_part}.xlsx"


def _prepare_sheet(worksheet) -> None:
    """Apply common worksheet settings."""
    worksheet.sheet_view.rightToLeft = True
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = "A1:Z1"


def _style_header(worksheet) -> None:
    """Style first row as a prominent report header."""
    for cell in worksheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _style_body_and_summary(worksheet, *, include_summary: bool = True) -> None:
    """Style body rows and optional final summary row."""
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = _BODY_FONT
            cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
    if include_summary and worksheet.max_row >= 2:
        for cell in worksheet[worksheet.max_row]:
            cell.fill = _SUMMARY_FILL
            cell.font = Font(name="Tahoma", bold=True)


def _set_widths(worksheet, widths: list[int]) -> None:
    """Set readable Excel column widths."""
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width


def _save_workbook(workbook: Workbook) -> BytesIO:
    """Save workbook to rewound BytesIO."""
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _format_dt(value) -> str:
    """Format datetime values for Excel cells."""
    if value is None:
        return ""
    return timezone.localtime(value).strftime("%Y-%m-%d %H:%M")
