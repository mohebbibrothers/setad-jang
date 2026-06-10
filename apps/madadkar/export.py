"""
Excel export engine اپ مددکار.

این ماژول یک فایل Excel حرفه‌ای برای گزارش‌گیری از پرداخت‌های یک حرکت
تولید می‌کند.

ویژگی‌ها:
- RTL alignment (راست به چپ — مناسب فارسی)
- Styled headers (پس‌زمینه، فونت bold، border، center)
- Summary row در پایین (مجموع سهم‌ها، مجموع مبلغ، تعداد مشارکت‌کنندگان)
- فرمت‌بندی اعداد با جداکننده هزارگان
- فرمت‌بندی تاریخ بر اساس timezone پروژه
- Auto-width برای ستون‌ها (با محاسبه طول محتوا)
- خروجی به‌صورت BytesIO که قابل ارسال به HttpResponse است
- نام‌گذاری sheet با عنوان حرکت (sanitize شده)

اصول طراحی:
- هیچ DB write در این ماژول انجام نمی‌شود — فقط read.
- خروجی in-memory است (BytesIO) — مناسب streaming.
- styling در constantها متمرکز شده تا قابل تغییر و reuse باشد.
"""

from __future__ import annotations

import io
import re
from datetime import datetime
from typing import TYPE_CHECKING

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

    from apps.madadkar.models import Campaign


# ===========================================================================
# Style constants
# ===========================================================================

_HEADER_FILL = PatternFill(
    start_color="1976D2",
    end_color="1976D2",
    fill_type="solid",
)
_HEADER_FONT = Font(
    name="Tahoma",
    size=11,
    bold=True,
    color="FFFFFF",
)
_HEADER_ALIGNMENT = Alignment(
    horizontal="center",
    vertical="center",
    wrap_text=True,
    readingOrder=2,  # RTL
)

_BODY_FONT = Font(name="Tahoma", size=10)
_BODY_ALIGNMENT_TEXT = Alignment(
    horizontal="right",
    vertical="center",
    readingOrder=2,
)
_BODY_ALIGNMENT_NUMBER = Alignment(
    horizontal="left",
    vertical="center",
)
_BODY_ALIGNMENT_CENTER = Alignment(
    horizontal="center",
    vertical="center",
)

_SUMMARY_FILL = PatternFill(
    start_color="E3F2FD",
    end_color="E3F2FD",
    fill_type="solid",
)
_SUMMARY_FONT = Font(name="Tahoma", size=11, bold=True)

_THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

_ROW_HEIGHT_HEADER = 32
_ROW_HEIGHT_BODY = 22
_ROW_HEIGHT_SUMMARY = 28

_NUMBER_FORMAT_INTEGER = "#,##0"
_NUMBER_FORMAT_DATE = "yyyy/mm/dd  HH:MM"

# ── Column definitions: (header, width, type)
_COLUMNS: list[tuple[str, int, str]] = [
    ("ردیف", 8, "int"),
    ("نام کاربر", 28, "text"),
    ("ایمیل", 32, "text"),
    ("شماره موبایل", 18, "text"),
    ("تعداد سهم", 14, "int"),
    ("قیمت سهم (تومان)", 22, "int"),
    ("مبلغ کل (تومان)", 22, "int"),
    ("کد رهگیری درگاه", 38, "text"),
    ("شناسه مرجع پرداخت", 28, "text"),
    ("نام درگاه", 14, "text"),
    ("تاریخ پرداخت", 22, "date"),
]


# ===========================================================================
# Helpers
# ===========================================================================

def _sanitize_sheet_name(name: str, max_length: int = 31) -> str:
    r"""
    پاکسازی نام sheet برای Excel.

    Excel محدودیت‌هایی برای نام sheet دارد:
    - حداکثر 31 کاراکتر
    - کاراکترهای ممنوع: : \ / ? * [ ]
    """
    cleaned = re.sub(r"[:\\/?*\[\]]", "-", name)
    cleaned = cleaned.strip()[:max_length]
    return cleaned or "Sheet1"


def _get_user_display_name(user) -> str:
    """نام نمایشی کاربر — fallback chain تا یک مقدار معتبر."""
    full_name = ""
    if hasattr(user, "get_full_name"):
        full_name = (user.get_full_name() or "").strip()
    if full_name:
        return full_name
    if hasattr(user, "first_name") and hasattr(user, "last_name"):
        combined = f"{user.first_name or ''} {user.last_name or ''}".strip()
        if combined:
            return combined
    return getattr(user, "email", "") or getattr(user, "username", "") or "—"


def _get_user_mobile(user) -> str:
    """شماره موبایل کاربر — defensive lookup."""
    return (
        getattr(user, "phone_number", "")
        or getattr(user, "mobile", "")
        or "—"
    )


def _localize_datetime(dt: datetime | None) -> datetime | str:
    """تبدیل datetime به timezone پروژه برای نمایش."""
    if dt is None:
        return "—"
    if timezone.is_aware(dt):
        return timezone.localtime(dt).replace(tzinfo=None)
    return dt


def _apply_header_styling(ws: Worksheet, row_idx: int) -> None:
    """اعمال styling به ردیف header."""
    for col_idx, _ in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _THIN_BORDER
    ws.row_dimensions[row_idx].height = _ROW_HEIGHT_HEADER


def _apply_column_widths(ws: Worksheet) -> None:
    """تنظیم عرض ستون‌ها."""
    for col_idx, (_header, width, _type) in enumerate(_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _apply_cell_styling(cell, column_type: str) -> None:
    """اعمال styling به یک سلول داده."""
    cell.font = _BODY_FONT
    cell.border = _THIN_BORDER

    if column_type == "int":
        cell.alignment = _BODY_ALIGNMENT_NUMBER
        cell.number_format = _NUMBER_FORMAT_INTEGER
    elif column_type == "date":
        cell.alignment = _BODY_ALIGNMENT_CENTER
        cell.number_format = _NUMBER_FORMAT_DATE
    else:
        cell.alignment = _BODY_ALIGNMENT_TEXT


# ===========================================================================
# Main export function
# ===========================================================================

def generate_campaign_participants_excel(*, campaign: Campaign) -> io.BytesIO:
    """
    تولید فایل Excel گزارش پرداخت‌های یک حرکت.

    این تابع تمام Participationهای PAID مربوط به campaign را به Excel
    تبدیل می‌کند. ترتیب: بزرگ‌ترین مبلغ ابتدا (descending by total_amount)،
    سپس بر اساس paid_at descending.

    Args:
        campaign: حرکتی که می‌خواهیم پرداخت‌هایش export شود.

    Returns:
        BytesIO حاوی فایل xlsx آماده برای ارسال در HttpResponse.

    نکات معماری:
    - این تابع import‌های selectors را داخل بدنه انجام می‌دهد تا
      circular import جلوگیری شود.
    - workbook در حافظه ساخته می‌شود — برای حرکت‌های بزرگ (>50k پرداخت)
      بهتر است streaming را در آینده اضافه کنیم.
    """
    # late import — جلوگیری از circular
    from apps.madadkar.selectors import get_campaign_participants_for_export

    participations = get_campaign_participants_for_export(campaign=campaign)

    wb = Workbook()
    ws = wb.active
    ws.sheet_view.rightToLeft = True
    ws.title = _sanitize_sheet_name(campaign.title)

    # ── Header row
    header_row = 1
    for col_idx, (header_label, _width, _type) in enumerate(
        _COLUMNS, start=1,
    ):
        ws.cell(row=header_row, column=col_idx, value=header_label)
    _apply_header_styling(ws, header_row)
    _apply_column_widths(ws)

    # ── Data rows
    total_shares = 0
    total_amount = 0
    unique_users: set[int] = set()

    current_row = header_row + 1
    for index, participation in enumerate(participations, start=1):
        user = participation.user
        payment = participation.payment

        values = [
            index,
            _get_user_display_name(user),
            getattr(user, "email", "") or "—",
            _get_user_mobile(user),
            participation.share_count,
            participation.share_price_snapshot,
            participation.total_amount,
            payment.authority if payment else "—",
            (payment.ref_id if payment and payment.ref_id else "—"),
            (payment.gateway_name if payment else "—"),
            _localize_datetime(participation.paid_at),
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            _apply_cell_styling(cell, _COLUMNS[col_idx - 1][2])

        ws.row_dimensions[current_row].height = _ROW_HEIGHT_BODY

        total_shares += participation.share_count
        total_amount += participation.total_amount
        unique_users.add(user.pk)

        current_row += 1

    # ── Summary row
    summary_row = current_row
    summary_label = (
        f"مجموع — {len(unique_users):,} مشارکت‌کننده یکتا"
    )
    ws.cell(row=summary_row, column=1, value=summary_label)
    ws.merge_cells(
        start_row=summary_row, start_column=1,
        end_row=summary_row, end_column=4,
    )
    ws.cell(row=summary_row, column=5, value=total_shares)
    ws.cell(row=summary_row, column=7, value=total_amount)

    for col_idx in range(1, len(_COLUMNS) + 1):
        cell = ws.cell(row=summary_row, column=col_idx)
        cell.fill = _SUMMARY_FILL
        cell.font = _SUMMARY_FONT
        cell.border = _THIN_BORDER
        if col_idx in (5, 7):
            cell.alignment = _BODY_ALIGNMENT_NUMBER
            cell.number_format = _NUMBER_FORMAT_INTEGER
        elif col_idx == 1:
            cell.alignment = _BODY_ALIGNMENT_CENTER
        else:
            cell.alignment = _BODY_ALIGNMENT_CENTER

    ws.row_dimensions[summary_row].height = _ROW_HEIGHT_SUMMARY

    # ── Freeze header
    ws.freeze_panes = "A2"

    # ── Output
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def build_excel_filename(*, campaign: Campaign) -> str:
    """
    ساخت نام فایل پیشنهادی برای دانلود.

    فرمت: madadkar-{campaign_id}-{slugified-title}-{YYYYMMDD-HHMMSS}.xlsx
    """
    now = timezone.localtime(timezone.now()).strftime("%Y%m%d-%H%M%S")
    safe_title = re.sub(r"[^\w-]+", "-", campaign.slug)[:60]
    return f"madadkar-{campaign.pk}-{safe_title}-{now}.xlsx"
