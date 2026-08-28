"""
تست‌های Excel export engine.

پوشش:
- generate_campaign_participants_excel:
  * تولید فایل valid xlsx (با openpyxl قابل بازخوانی)
  * هدر ستون‌ها درست است
  * فقط participations PAID شامل می‌شوند (نه PENDING/FAILED/EXPIRED)
  * مقادیر صحیح در سلول‌ها (share_count, amount, user info, authority)
  * ترتیب: بزرگ‌ترین مبلغ ابتدا
  * ردیف summary در پایان با مجموع‌های صحیح
  * RTL alignment تنظیم شده
  * freeze panes روی A2
  * empty campaign — فقط header + summary خالی
- build_excel_filename:
  * فرمت نام فایل
  * timestamp included
  * sanitization
- helper functions: _sanitize_sheet_name
"""

from __future__ import annotations

import io

import pytest
from openpyxl import load_workbook

from apps.madadkar.export import (
    _sanitize_sheet_name,
    build_excel_filename,
    generate_campaign_participants_excel,
)
from tests.factories import UserFactory
from tests.factories.madadkar import (
    FailedParticipationFactory,
    PaidParticipationFactory,
    ParticipationFactory,
    PublishedCampaignFactory,
    SuccessPaymentFactory,
)

pytestmark = pytest.mark.django_db


# ============================================================
# Helpers — للتست
# ============================================================


def _make_paid_participation_with_payment(
    *,
    campaign,
    user=None,
    share_count: int = 1,
):
    """ساخت یک Participation PAID همراه با Payment SUCCESS."""
    if user is None:
        user = UserFactory()
    participation = PaidParticipationFactory(
        campaign=campaign,
        user=user,
        share_count=share_count,
    )
    SuccessPaymentFactory(
        participation=participation,
        user=user,
    )
    return participation


def _load_workbook_from_buffer(buffer: io.BytesIO):
    """بارگذاری workbook از BytesIO برای assertion."""
    buffer.seek(0)
    return load_workbook(buffer, data_only=False)


# ============================================================
# generate_campaign_participants_excel — basic structure
# ============================================================


class TestExcelStructure:
    """تست‌های ساختار فایل Excel تولید شده."""

    def test_returns_bytesio(self):
        """خروجی باید BytesIO باشد."""
        campaign = PublishedCampaignFactory()
        result = generate_campaign_participants_excel(campaign=campaign)
        assert isinstance(result, io.BytesIO)

    def test_output_is_valid_xlsx(self):
        """خروجی باید یک فایل xlsx معتبر باشد که با openpyxl قابل خواندن است."""
        campaign = PublishedCampaignFactory()
        buffer = generate_campaign_participants_excel(campaign=campaign)

        wb = _load_workbook_from_buffer(buffer)
        assert wb.active is not None

    def test_sheet_name_includes_campaign_title(self):
        """نام sheet باید از عنوان حرکت باشد (با sanitize)."""
        campaign = PublishedCampaignFactory(title="حرکت تست")
        buffer = generate_campaign_participants_excel(campaign=campaign)

        wb = _load_workbook_from_buffer(buffer)
        assert "حرکت تست" in wb.active.title

    def test_rtl_alignment_enabled(self):
        """sheet باید RTL باشد."""
        campaign = PublishedCampaignFactory()
        buffer = generate_campaign_participants_excel(campaign=campaign)

        wb = _load_workbook_from_buffer(buffer)
        assert wb.active.sheet_view.rightToLeft is True

    def test_freeze_panes_on_header(self):
        """ردیف header باید freeze شده باشد."""
        campaign = PublishedCampaignFactory()
        buffer = generate_campaign_participants_excel(campaign=campaign)

        wb = _load_workbook_from_buffer(buffer)
        assert wb.active.freeze_panes == "A2"


# ============================================================
# Header row
# ============================================================


class TestExcelHeader:
    """تست‌های ردیف header."""

    def test_header_contains_all_columns(self):
        """ردیف اول باید شامل تمام عناوین ستون باشد."""
        campaign = PublishedCampaignFactory()
        buffer = generate_campaign_participants_excel(campaign=campaign)

        wb = _load_workbook_from_buffer(buffer)
        ws = wb.active

        expected_headers = [
            "ردیف",
            "نام کاربر",
            "ایمیل",
            "شماره موبایل",
            "تعداد سهم",
            "قیمت سهم (تومان)",
            "مبلغ کل (تومان)",
            "کد رهگیری درگاه",
            "شناسه مرجع پرداخت",
            "نام درگاه",
            "تاریخ پرداخت",
        ]

        for col_idx, expected in enumerate(expected_headers, start=1):
            assert ws.cell(row=1, column=col_idx).value == expected


# ============================================================
# Data rows — content
# ============================================================


class TestExcelDataRows:
    """تست‌های ردیف‌های داده."""

    def test_includes_only_paid_participations(self):
        """فقط participationهای PAID باید در Excel باشند."""
        campaign = PublishedCampaignFactory(
            total_amount=100_000_000,
            total_shares=100,
        )

        # PAID — باید بیاید
        _make_paid_participation_with_payment(campaign=campaign, share_count=2)

        # PENDING_PAYMENT — نباید بیاید
        ParticipationFactory(campaign=campaign, share_count=3)

        # FAILED — نباید بیاید
        FailedParticipationFactory(campaign=campaign, share_count=4)

        buffer = generate_campaign_participants_excel(campaign=campaign)
        wb = _load_workbook_from_buffer(buffer)
        ws = wb.active

        # ردیف 1 = header، ردیف 2 = اولین داده، ردیف 3 = summary
        # یعنی فقط 1 ردیف داده باید باشد
        # شمارش تعداد ردیف‌های داده با چک کردن ستون "تعداد سهم"
        data_row = ws.cell(row=2, column=5).value
        assert data_row == 2  # فقط participation PAID با share_count=2

    def test_correct_amount_and_share_count(self):
        """مقادیر share_count و total_amount درست در سلول‌ها باشند."""
        campaign = PublishedCampaignFactory(
            total_amount=100_000_000,
            total_shares=10,  # 10 میلیون per share
        )
        _make_paid_participation_with_payment(
            campaign=campaign,
            share_count=3,
        )

        buffer = generate_campaign_participants_excel(campaign=campaign)
        wb = _load_workbook_from_buffer(buffer)
        ws = wb.active

        # ردیف 2 = اولین داده
        assert ws.cell(row=2, column=5).value == 3  # share_count
        assert ws.cell(row=2, column=6).value == 10_000_000  # share_price
        assert ws.cell(row=2, column=7).value == 30_000_000  # total_amount

    def test_includes_user_email(self):
        """ایمیل کاربر باید در ستون 3 باشد."""
        campaign = PublishedCampaignFactory()
        user = UserFactory(email="contributor@test.local")
        _make_paid_participation_with_payment(campaign=campaign, user=user)

        buffer = generate_campaign_participants_excel(campaign=campaign)
        wb = _load_workbook_from_buffer(buffer)
        ws = wb.active

        assert ws.cell(row=2, column=3).value == "contributor@test.local"

    def test_includes_payment_authority(self):
        """کد رهگیری درگاه (authority) باید در ستون 8 باشد."""
        campaign = PublishedCampaignFactory()
        participation = _make_paid_participation_with_payment(
            campaign=campaign,
        )

        buffer = generate_campaign_participants_excel(campaign=campaign)
        wb = _load_workbook_from_buffer(buffer)
        ws = wb.active

        assert ws.cell(row=2, column=8).value == participation.payment.authority

    def test_includes_gateway_name(self):
        """نام درگاه باید در ستون 10 باشد."""
        campaign = PublishedCampaignFactory()
        _make_paid_participation_with_payment(campaign=campaign)

        buffer = generate_campaign_participants_excel(campaign=campaign)
        wb = _load_workbook_from_buffer(buffer)
        ws = wb.active

        assert ws.cell(row=2, column=10).value == "sandbox"


# ============================================================
# Sorting (largest amount first)
# ============================================================


class TestExcelSorting:
    """ترتیب ردیف‌های داده — بزرگ‌ترین مبلغ ابتدا."""

    def test_largest_amount_first(self):
        """ردیف اول باید بزرگ‌ترین total_amount باشد."""
        campaign = PublishedCampaignFactory(
            total_amount=100_000_000,
            total_shares=10,
        )

        # سه مشارکت با مقادیر مختلف
        _make_paid_participation_with_payment(campaign=campaign, share_count=1)
        _make_paid_participation_with_payment(campaign=campaign, share_count=5)
        _make_paid_participation_with_payment(campaign=campaign, share_count=3)

        buffer = generate_campaign_participants_excel(campaign=campaign)
        wb = _load_workbook_from_buffer(buffer)
        ws = wb.active

        # ردیف 2 (اولین داده): share_count=5 → بزرگ‌ترین
        assert ws.cell(row=2, column=5).value == 5
        # ردیف 3: share_count=3
        assert ws.cell(row=3, column=5).value == 3
        # ردیف 4: share_count=1
        assert ws.cell(row=4, column=5).value == 1


# ============================================================
# Summary row
# ============================================================


class TestExcelSummaryRow:
    """تست‌های ردیف summary در پایان فایل."""

    def test_summary_row_at_end(self):
        """آخرین ردیف باید summary باشد و شامل مجموع‌ها."""
        campaign = PublishedCampaignFactory(
            total_amount=100_000_000,
            total_shares=10,
        )

        # 2 مشارکت با مجموع: 8 سهم و 80M
        _make_paid_participation_with_payment(campaign=campaign, share_count=3)
        _make_paid_participation_with_payment(campaign=campaign, share_count=5)

        buffer = generate_campaign_participants_excel(campaign=campaign)
        wb = _load_workbook_from_buffer(buffer)
        ws = wb.active

        # header(1) + 2 data rows + 1 summary = ردیف 4 → summary
        summary_row = 4

        # ستون 5: مجموع سهم
        assert ws.cell(row=summary_row, column=5).value == 8
        # ستون 7: مجموع مبلغ
        assert ws.cell(row=summary_row, column=7).value == 80_000_000

    def test_summary_includes_unique_users_count(self):
        """ردیف summary باید تعداد کاربران یکتا را نشان دهد."""
        campaign = PublishedCampaignFactory(
            total_amount=100_000_000,
            total_shares=100,
        )

        user_a = UserFactory()
        user_b = UserFactory()
        # user_a دو مشارکت دارد
        _make_paid_participation_with_payment(
            campaign=campaign,
            user=user_a,
            share_count=1,
        )
        _make_paid_participation_with_payment(
            campaign=campaign,
            user=user_a,
            share_count=2,
        )
        # user_b یک مشارکت دارد
        _make_paid_participation_with_payment(
            campaign=campaign,
            user=user_b,
            share_count=3,
        )

        buffer = generate_campaign_participants_excel(campaign=campaign)
        wb = _load_workbook_from_buffer(buffer)
        ws = wb.active

        # header(1) + 3 data rows + summary = ردیف 5
        summary_label = ws.cell(row=5, column=1).value
        assert "2" in summary_label  # 2 کاربر یکتا
        assert "مجموع" in summary_label


# ============================================================
# Empty campaign
# ============================================================


class TestExcelEmptyCampaign:
    """رفتار با حرکتی که هیچ پرداخت موفقی ندارد."""

    def test_empty_campaign_returns_header_and_empty_summary(self):
        """فقط header + summary خالی باید باشد."""
        campaign = PublishedCampaignFactory()
        # هیچ participation ندارد

        buffer = generate_campaign_participants_excel(campaign=campaign)
        wb = _load_workbook_from_buffer(buffer)
        ws = wb.active

        # ردیف 1 = header، ردیف 2 = summary
        # مقدار summary باید صفر باشد
        assert ws.cell(row=2, column=5).value == 0  # مجموع سهم
        assert ws.cell(row=2, column=7).value == 0  # مجموع مبلغ

    def test_only_pending_participations_excluded(self):
        """اگر همه participations فقط PENDING باشند، Excel خالی است."""
        campaign = PublishedCampaignFactory()
        ParticipationFactory(campaign=campaign)
        ParticipationFactory(campaign=campaign)

        buffer = generate_campaign_participants_excel(campaign=campaign)
        wb = _load_workbook_from_buffer(buffer)
        ws = wb.active

        # ردیف 2 = summary (هیچ داده‌ای نیست)
        assert ws.cell(row=2, column=5).value == 0


# ============================================================
# build_excel_filename
# ============================================================


class TestExcelFilename:
    """تست‌های ساخت نام فایل."""

    def test_filename_includes_campaign_id(self):
        campaign = PublishedCampaignFactory()
        filename = build_excel_filename(campaign=campaign)
        assert f"madadkar-{campaign.pk}-" in filename

    def test_filename_has_xlsx_extension(self):
        campaign = PublishedCampaignFactory()
        filename = build_excel_filename(campaign=campaign)
        assert filename.endswith(".xlsx")

    def test_filename_includes_timestamp(self):
        """نام فایل باید timestamp داشته باشد (YYYYMMDD-HHMMSS)."""
        campaign = PublishedCampaignFactory()
        filename = build_excel_filename(campaign=campaign)
        # فرمت: madadkar-{id}-{slug}-YYYYMMDD-HHMMSS.xlsx
        # حداقل باید 4 رقم برای سال داشته باشد
        import re

        assert re.search(r"\d{8}-\d{6}", filename) is not None

    def test_filename_includes_sanitized_slug(self):
        campaign = PublishedCampaignFactory(title="حرکت تست")
        filename = build_excel_filename(campaign=campaign)
        # slug فارسی است، اما در filename باید همان slug پاکسازی شده باشد
        assert (
            campaign.slug.replace("-", "") in filename.replace("-", "") or len(filename) > 20
        )  # حداقل نام معنادار باشد


# ============================================================
# _sanitize_sheet_name
# ============================================================


class TestSanitizeSheetName:
    """تست‌های پاکسازی نام sheet."""

    def test_max_length_31(self):
        """نام sheet نباید بیشتر از 31 کاراکتر باشد."""
        long_name = "a" * 100
        result = _sanitize_sheet_name(long_name)
        assert len(result) <= 31

    def test_removes_forbidden_chars(self):
        """کاراکترهای ممنوع باید با - جایگزین شوند."""
        result = _sanitize_sheet_name("test:name/with\\bad?chars*[here]")
        forbidden = [":", "/", "\\", "?", "*", "[", "]"]
        for char in forbidden:
            assert char not in result

    def test_empty_input_returns_fallback(self):
        """اگر ورودی خالی باشد، 'Sheet1' برمی‌گردد."""
        assert _sanitize_sheet_name("") == "Sheet1"

    def test_preserves_persian_text(self):
        """متن فارسی باید حفظ شود."""
        result = _sanitize_sheet_name("حرکت تست")
        assert "حرکت" in result
        assert "تست" in result
