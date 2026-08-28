"""Settlement import/export helpers for Madadkar payment reconciliation."""

from __future__ import annotations

import csv
from io import BytesIO, StringIO
from typing import Any

from openpyxl import load_workbook

from apps.madadkar.choices import ReconciliationItemStatus
from apps.madadkar.models import PaymentReconciliationBatch

RECONCILIATION_REQUIRED_COLUMNS = {"amount", "status"}
RECONCILIATION_IDENTITY_COLUMNS = {"authority", "ref_id"}
_HEADER_ALIASES = {
    "authority": "authority",
    "Authority": "authority",
    "authority_code": "authority",
    "ref_id": "ref_id",
    "refid": "ref_id",
    "reference": "ref_id",
    "reference_id": "ref_id",
    "amount": "amount",
    "Amount": "amount",
    "مبلغ": "amount",
    "status": "status",
    "Status": "status",
    "وضعیت": "status",
}


class ReconciliationImportError(ValueError):
    """Raised when a settlement report cannot be parsed safely."""


def parse_settlement_report(*, filename: str, content: bytes) -> list[dict[str, Any]]:
    """Parse CSV/XLSX provider settlement rows into canonical reconciliation rows."""
    normalized_name = filename.lower().strip()
    if normalized_name.endswith(".csv"):
        rows = _parse_csv(content=content)
    elif normalized_name.endswith(".xlsx"):
        rows = _parse_xlsx(content=content)
    else:
        raise ReconciliationImportError("فرمت فایل تطبیق باید CSV یا XLSX باشد.")
    if not rows:
        raise ReconciliationImportError("فایل تطبیق هیچ ردیف معتبری ندارد.")
    return rows


def build_reconciliation_discrepancy_csv(*, batch: PaymentReconciliationBatch) -> bytes:
    """Render non-matched reconciliation items as UTF-8-SIG CSV for finance review."""
    output = StringIO()
    fieldnames = [
        "item_id",
        "status",
        "reason",
        "authority",
        "provider_ref_id",
        "provider_amount",
        "provider_status",
        "internal_amount",
        "internal_status",
        "payment_id",
    ]
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    for item in batch.items.exclude(status=ReconciliationItemStatus.MATCHED).order_by(
        "created_at", "id"
    ):
        writer.writerow(
            {
                "item_id": item.pk,
                "status": item.status,
                "reason": item.reason,
                "authority": item.authority,
                "provider_ref_id": item.provider_ref_id,
                "provider_amount": item.provider_amount,
                "provider_status": item.provider_status,
                "internal_amount": item.internal_amount,
                "internal_status": item.internal_status,
                "payment_id": item.payment_id,
            }
        )
    return output.getvalue().encode("utf-8-sig")


def _parse_csv(*, content: bytes) -> list[dict[str, Any]]:
    """Parse CSV bytes with BOM support and canonical header mapping."""
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(StringIO(text))
    if reader.fieldnames is None:
        raise ReconciliationImportError("فایل CSV header ندارد.")
    headers = _normalize_headers(reader.fieldnames)
    rows = []
    for raw_row in reader:
        rows.append(_canonicalize_row(raw_row=raw_row, headers=headers))
    return _validate_rows(rows)


def _parse_xlsx(*, content: bytes) -> list[dict[str, Any]]:
    """Parse first XLSX worksheet into canonical settlement rows."""
    workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    raw_headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    headers = _normalize_headers([str(value or "") for value in raw_headers])
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        raw_row = {header: value for header, value in zip(raw_headers, row, strict=False)}
        rows.append(_canonicalize_row(raw_row=raw_row, headers=headers))
    return _validate_rows(rows)


def _normalize_headers(fieldnames: list[str]) -> dict[str, str]:
    """Map provider-specific headers to canonical reconciliation keys."""
    mapping = {}
    for header in fieldnames:
        stripped = str(header or "").strip()
        canonical = _HEADER_ALIASES.get(stripped, stripped.lower())
        mapping[stripped] = canonical
    canonical_values = set(mapping.values())
    missing = RECONCILIATION_REQUIRED_COLUMNS - canonical_values
    has_identity = bool(RECONCILIATION_IDENTITY_COLUMNS & canonical_values)
    if missing or not has_identity:
        raise ReconciliationImportError("ستون‌های ضروری فایل تطبیق کامل نیستند.")
    return mapping


def _canonicalize_row(*, raw_row: dict[Any, Any], headers: dict[str, str]) -> dict[str, Any]:
    """Convert one provider row into canonical authority/ref_id/amount/status fields."""
    canonical: dict[str, Any] = {"authority": "", "ref_id": "", "amount": 0, "status": ""}
    for raw_key, value in raw_row.items():
        key = headers.get(str(raw_key or "").strip())
        if key in canonical:
            canonical[key] = _normalize_cell(value=value, key=key)
    canonical["raw_payload"] = {str(key): value for key, value in raw_row.items()}
    return canonical


def _normalize_cell(*, value: Any, key: str) -> Any:
    """Normalize settlement cell values while preserving raw payload separately."""
    if key == "amount":
        try:
            return int(str(value or "0").replace(",", "").strip())
        except ValueError as exc:
            raise ReconciliationImportError("ستون amount باید عددی باشد.") from exc
    return str(value or "").strip()


def _validate_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Validate parsed rows and strip fully empty rows."""
    cleaned = [
        row
        for row in rows
        if row.get("authority") or row.get("ref_id") or row.get("amount") or row.get("status")
    ]
    for row in cleaned:
        if not row.get("authority") and not row.get("ref_id"):
            raise ReconciliationImportError("هر ردیف باید authority یا ref_id داشته باشد.")
        if int(row.get("amount") or 0) <= 0:
            raise ReconciliationImportError("مبلغ هر ردیف باید بزرگ‌تر از صفر باشد.")
        if not row.get("status"):
            raise ReconciliationImportError("status هر ردیف الزامی است.")
    return cleaned
