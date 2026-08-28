"""Streaming Excel writer shared by every admin export in the project.

چرا این ماژول وجود دارد
=======================
هر اپ export اکسل خودش را با ``openpyxl.Workbook()`` معمولی می‌ساخت. در آن
حالت openpyxl **کل workbook را به‌صورت شیء در حافظه نگه می‌دارد**: برای هر
سلول یک آبجکت پایتون با استایل، فرمت و مقدار. برای یک حرکت با ۵۰٬۰۰۰
مشارکت‌کننده و ۱۱ ستون یعنی بیش از نیم میلیون آبجکت — چند صد مگابایت RAM
داخل یک worker گانیکورن که همان مدت هم مسدود می‌ماند و می‌تواند با OOM
کشته شود.

بدتر اینکه الگوی رایج در آن ماژول‌ها این بود که اول همهٔ ردیف‌ها append
شوند و بعد با ``iter_rows`` یک پاس دیگر برای استایل‌دهی روی *همهٔ* سلول‌ها
زده شود. یعنی دو برابر پیمایش و نگه‌داری اجباری کل شبکه در حافظه.

راهکار
=======
``Workbook(write_only=True)`` هر ردیف را بلافاصله پس از append در قالب XML
روی فایل موقت می‌نویسد و از حافظه بیرون می‌اندازد. مصرف حافظه به‌جای
«متناسب با تعداد ردیف» می‌شود «متناسب با یک ردیف». استایل روی
``WriteOnlyCell`` همان‌جا هنگام نوشتن اعمال می‌شود، پس پاس دوم حذف می‌شود.

محدودیت پذیرفته‌شده: در حالت write-only نمی‌توان سلول ادغام‌شده
(``merge_cells``) ساخت و نمی‌توان بعد از نوشتن به سلول برگشت. ردیف خلاصه
به‌جای برچسب ادغام‌شده، برچسب را در ستون اول می‌نویسد. این تنها تفاوت
دیداری با خروجی قبلی است.

سقف ردیف
=========
حتی با نوشتن جریانی، یک export بی‌کران می‌تواند فایلی بسازد که نه دانلود
می‌شود و نه باز. ``max_rows`` یک سقف صریح می‌گذارد و در صورت عبور خطای
روشن می‌دهد، به‌جای اینکه worker بی‌صدا دقایقی مشغول بماند.
"""

from __future__ import annotations

from collections.abc import Iterable, Sequence
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from typing import Any

from django.conf import settings
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

#: سقف پیش‌فرض تعداد ردیف داده در یک export.
DEFAULT_MAX_EXPORT_ROWS = 200_000

_THIN_SIDE = Side(style="thin", color="D0D0D0")
_THIN_BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE)

_ALIGN_RIGHT = Alignment(horizontal="right", vertical="center", readingOrder=2)
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
_ALIGN_NUMBER = Alignment(horizontal="left", vertical="center")
_ALIGN_HEADER = Alignment(horizontal="center", vertical="center", wrap_text=True, readingOrder=2)

_NUMBER_FORMAT_INTEGER = "#,##0"
_NUMBER_FORMAT_DECIMAL = "#,##0.00"
_NUMBER_FORMAT_DATE = "yyyy-mm-dd hh:mm"


class ExcelExportTooLargeError(Exception):
    """Raised when an export exceeds the configured maximum row count."""


@dataclass(frozen=True)
class ExcelColumn:
    """One export column: header label, display width and value kind.

    ``kind`` تعیین می‌کند سلول چطور تراز و فرمت شود:
    ``text`` (راست‌چین)، ``int``/``decimal`` (چپ‌چین با جداکنندهٔ هزارگان)،
    ``date`` (وسط‌چین با فرمت تاریخ) و ``center`` (وسط‌چین ساده).
    """

    header: str
    width: int = 18
    kind: str = "text"


@dataclass(frozen=True)
class ExcelTheme:
    """Colour and font theme for one export sheet."""

    header_color: str = "1976D2"
    header_font_color: str = "FFFFFF"
    summary_color: str = "E3F2FD"
    font_name: str = "Tahoma"
    header_font_size: int = 11
    body_font_size: int = 10


def get_max_export_rows() -> int:
    """Return the configured hard cap on exported data rows."""
    return int(getattr(settings, "EXPORT_MAX_ROWS", DEFAULT_MAX_EXPORT_ROWS))


def localize_for_excel(value: datetime | date | None, *, empty: str = "—") -> Any:
    """Convert a datetime to project-local naive time for an Excel cell.

    اکسل مفهوم timezone ندارد؛ اگر datetime آگاه از منطقهٔ زمانی مستقیم نوشته
    شود openpyxl خطا می‌دهد. اینجا به وقت محلی پروژه تبدیل و naive می‌شود تا
    عددی که کاربر می‌بیند همان چیزی باشد که در پنل می‌بیند.
    """
    if value is None:
        return empty
    if isinstance(value, datetime) and timezone.is_aware(value):
        return timezone.localtime(value).replace(tzinfo=None)
    return value


def stream_rows(source: Any, *, chunk_size: int = 2_000) -> Iterable[Any]:
    """Iterate a queryset without loading it entirely into memory.

    اگر ``source`` یک queryset باشد از ``iterator(chunk_size=...)`` استفاده
    می‌شود تا نتیجهٔ کامل در ``_result_cache`` انباشته نشود؛ در غیر این صورت
    (لیست، تاپل، ژنراتور) همان‌طور که هست پیمایش می‌شود. این کار به
    توابع export اجازه می‌دهد بدون تغییر امضای عمومی‌شان جریانی شوند.
    """
    iterator = getattr(source, "iterator", None)
    if callable(iterator):
        return iterator(chunk_size=chunk_size)
    return source


def sanitize_sheet_title(value: str, *, fallback: str = "گزارش") -> str:
    """Return an Excel-safe worksheet title.

    اکسل نام sheet را حداکثر ۳۱ کاراکتر می‌پذیرد و کاراکترهای
    ``: \\ / ? * [ ]`` را ممنوع کرده است.
    """
    forbidden = set(":\\/?*[]")
    cleaned = "".join("-" if char in forbidden else char for char in (value or "")).strip()
    return cleaned[:31] or fallback


class StreamingExcelSheet:
    """A single write-only, RTL, styled worksheet built row by row.

    استفادهٔ معمول::

        sheet = StreamingExcelSheet(title="گزارش", columns=[...])
        for row in queryset.iterator(chunk_size=2000):
            sheet.append([...])
        sheet.append_summary(["مجموع", None, 42])
        buffer = sheet.save()

    هر ``append`` بلافاصله روی دیسک نوشته می‌شود، پس مصرف حافظه مستقل از
    تعداد ردیف‌ها ثابت می‌ماند.
    """

    def __init__(
        self,
        *,
        title: str,
        columns: Sequence[ExcelColumn],
        rtl: bool = True,
        freeze_header: bool = True,
        auto_filter: bool = False,
        theme: ExcelTheme | None = None,
        max_rows: int | None = None,
    ) -> None:
        self._columns = list(columns)
        self._theme = theme or ExcelTheme()
        self._max_rows = get_max_export_rows() if max_rows is None else max_rows
        self._row_count = 0
        self._saved = False

        self._workbook = Workbook(write_only=True)
        self._sheet = self._workbook.create_sheet(title=sanitize_sheet_title(title))
        self._sheet.sheet_view.rightToLeft = rtl

        # عرض ستون‌ها و freeze باید *قبل* از نوشتن اولین ردیف تنظیم شوند،
        # چون در حالت write-only بعد از شروع نوشتن دیگر قابل تغییر نیستند.
        for index, column in enumerate(self._columns, start=1):
            self._sheet.column_dimensions[get_column_letter(index)].width = column.width
        if freeze_header:
            self._sheet.freeze_panes = "A2"
        if auto_filter and self._columns:
            self._sheet.auto_filter.ref = f"A1:{get_column_letter(len(self._columns))}1"

        self._write_header()

    @property
    def row_count(self) -> int:
        """Number of data rows written so far, excluding header and summary."""
        return self._row_count

    def _styled_cell(
        self, value: Any, *, kind: str, header: bool = False, summary: bool = False
    ) -> Any:
        """Build one styled write-only cell.

        datetimeهای آگاه از منطقهٔ زمانی همین‌جا به وقت محلی naive تبدیل
        می‌شوند. اکسل tzinfo را نمی‌پذیرد و openpyxl در آن حالت ``TypeError``
        می‌دهد؛ انجام این کار در یک نقطهٔ مرکزی جلوی تکرار همان باگ در هر
        فراخوانندهٔ جدید را می‌گیرد.
        """
        if isinstance(value, datetime) and timezone.is_aware(value):
            value = timezone.localtime(value).replace(tzinfo=None)
        cell = WriteOnlyCell(self._sheet, value=value)
        theme = self._theme
        if header:
            cell.fill = PatternFill(
                start_color=theme.header_color, end_color=theme.header_color, fill_type="solid"
            )
            cell.font = Font(
                name=theme.font_name,
                size=theme.header_font_size,
                bold=True,
                color=theme.header_font_color,
            )
            cell.alignment = _ALIGN_HEADER
            cell.border = _THIN_BORDER
            return cell

        cell.font = Font(name=theme.font_name, size=theme.body_font_size, bold=summary)
        cell.border = _THIN_BORDER
        if summary:
            cell.fill = PatternFill(
                start_color=theme.summary_color, end_color=theme.summary_color, fill_type="solid"
            )

        if kind == "int":
            cell.alignment = _ALIGN_NUMBER
            cell.number_format = _NUMBER_FORMAT_INTEGER
        elif kind == "decimal":
            cell.alignment = _ALIGN_NUMBER
            cell.number_format = _NUMBER_FORMAT_DECIMAL
        elif kind == "date":
            cell.alignment = _ALIGN_CENTER
            cell.number_format = _NUMBER_FORMAT_DATE
        elif kind == "center":
            cell.alignment = _ALIGN_CENTER
        else:
            cell.alignment = _ALIGN_RIGHT
        return cell

    def _write_header(self) -> None:
        """Write the styled header row."""
        self._sheet.append(
            [
                self._styled_cell(column.header, kind="center", header=True)
                for column in self._columns
            ],
        )

    def append(self, values: Sequence[Any]) -> None:
        """Append one styled data row.

        Raises:
            ExcelExportTooLargeError: اگر تعداد ردیف‌ها از سقف عبور کند.
        """
        self._row_count += 1
        if self._row_count > self._max_rows:
            # قبل از پرتاب خطا workbook را می‌بندیم. در حالت write-only یک
            # فایل موقت و یک ژنراتور باز وجود دارد؛ رها کردنشان یعنی نشت
            # منبع تا زمان GC، دقیقاً در سناریویی که گزارش خیلی بزرگ بوده.
            self.close()
            raise ExcelExportTooLargeError(
                f"تعداد ردیف‌های این گزارش از سقف مجاز ({self._max_rows:,}) بیشتر است. "
                "بازهٔ گزارش را محدودتر کنید.",
            )
        self._sheet.append(
            [
                self._styled_cell(value, kind=column.kind)
                for value, column in zip(values, self._columns, strict=False)
            ],
        )

    def extend(self, rows: Iterable[Sequence[Any]]) -> None:
        """Append many rows, keeping memory flat."""
        for row in rows:
            self.append(row)

    def append_summary(self, values: Sequence[Any]) -> None:
        """Append a highlighted summary row.

        در حالت write-only ادغام سلول ممکن نیست، پس برچسب خلاصه در ستون اول
        نوشته می‌شود و ستون‌های بدون مقدار خالی می‌مانند.
        """
        self._sheet.append(
            [
                self._styled_cell(
                    value, kind=column.kind if value is not None else "center", summary=True
                )
                for value, column in zip(values, self._columns, strict=False)
            ],
        )

    def save(self) -> BytesIO:
        """Finalize the workbook and return it as a seek-to-zero buffer."""
        if self._saved:
            raise RuntimeError("این workbook قبلاً ذخیره شده است.")
        self._saved = True
        buffer = BytesIO()
        self._workbook.save(buffer)
        buffer.seek(0)
        return buffer

    def close(self) -> None:
        """Release the write-only temporary file without producing output.

        در حالت write-only هر sheet یک ژنراتور ردیف و یک فایل موقت باز دارد.
        اگر export نیمه‌کاره رها شود (خطای سقف ردیف، استثنا در حلقهٔ داده)
        باید صریحاً بسته شود، وگرنه منبع تا اجرای garbage collector باز
        می‌ماند و در آن نقطه هم با «I/O operation on closed file» می‌ترکد.

        نکتهٔ ظریف: ``Workbook.close()`` در این حالت فقط آرشیو zip را می‌بندد
        که هنوز ساخته نشده. چیزی که واقعاً منبع را آزاد می‌کند
        ``WriteOnlyWorksheet.close()`` است. فراخوانی دوباره بی‌اثر است.
        """
        if self._saved:
            return
        self._saved = True
        if not self._sheet.closed:
            self._sheet.close()
        self._workbook.close()

    def __enter__(self) -> StreamingExcelSheet:
        """Enter the writer context."""
        return self

    def __exit__(self, exc_type, exc_value, traceback) -> None:
        """Close the workbook if the body exited without saving."""
        self.close()
