"""Excel export helpers for LMS admin analytics.

Exports are generated in-memory with openpyxl and are intentionally kept outside
views/services so reporting output remains testable and replaceable.
"""

from __future__ import annotations

from collections.abc import Iterable
from io import BytesIO

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from apps.lms.models import Course, Enrollment

_HEADER_FILL = PatternFill("solid", fgColor="277A7E")
_HEADER_FONT = Font(name="Tahoma", bold=True, color="FFFFFF")
_BODY_FONT = Font(name="Tahoma", size=11)


def build_course_enrollments_workbook(*, course: Course, enrollments: Iterable[Enrollment]) -> BytesIO:
    """Build an RTL Excel workbook for a course participant report."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = _safe_sheet_title(course.title)
    worksheet.sheet_view.rightToLeft = True
    worksheet.freeze_panes = "A2"

    headers = [
        "شناسه ثبت‌نام",
        "نام کامل",
        "ایمیل",
        "وضعیت",
        "درصد پیشرفت",
        "ثانیه مشاهده‌شده",
        "کد مدرک",
        "تاریخ ثبت‌نام",
        "تاریخ تکمیل",
    ]
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for enrollment in enrollments:
        user = enrollment.user
        certificate = getattr(enrollment, "certificate", None)
        worksheet.append(
            [
                enrollment.pk,
                getattr(user, "full_name", "") or str(user),
                getattr(user, "email", "") or "",
                enrollment.get_status_display(),
                float(enrollment.progress_percent or 0),
                enrollment.watched_seconds,
                getattr(certificate, "certificate_code", "") if certificate else "",
                _format_dt(enrollment.enrolled_at),
                _format_dt(enrollment.completed_at),
            ]
        )

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = _BODY_FONT
            cell.alignment = Alignment(horizontal="right")

    for index, width in enumerate([16, 28, 32, 18, 18, 20, 28, 24, 24], start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def build_course_export_filename(*, course: Course) -> str:
    """Build a deterministic export filename for a course report."""
    date_part = timezone.now().strftime("%Y%m%d")
    return f"lms-course-{course.pk}-participants-{date_part}.xlsx"


def _format_dt(value) -> str:
    """Format datetime values for Excel cells."""
    if value is None:
        return ""
    return timezone.localtime(value).strftime("%Y-%m-%d %H:%M")


def _safe_sheet_title(value: str) -> str:
    """Return an Excel-safe sheet title."""
    forbidden = set('[]:*?/\\')
    cleaned = "".join("-" if char in forbidden else char for char in value).strip() or "گزارش"
    return cleaned[:31]
