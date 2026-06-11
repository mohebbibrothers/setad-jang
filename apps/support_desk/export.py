"""Excel export helpers for Support Desk admin reporting."""

from __future__ import annotations

from collections.abc import Iterable
from io import BytesIO

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from apps.support_desk.models import SupportTicket, SupportTicketMessage, SupportTicketSatisfaction

_HEADER_FILL = PatternFill("solid", fgColor="4B2E83")
_HEADER_FONT = Font(name="Tahoma", bold=True, color="FFFFFF")
_BODY_FONT = Font(name="Tahoma", size=11)
_SUMMARY_FILL = PatternFill("solid", fgColor="EFE9FF")


def build_tickets_workbook(*, tickets: Iterable[SupportTicket]) -> BytesIO:
    """Build an RTL Excel workbook for support ticket queue/export."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "تیکت‌ها"
    _prepare_sheet(worksheet)
    headers = [
        "شماره تیکت",
        "موضوع",
        "مالک",
        "دپارتمان",
        "دسته",
        "نوع",
        "وضعیت",
        "اولویت",
        "شدت",
        "مسئول",
        "SLA نقض شده؟",
        "تعداد پیام",
        "تعداد ضمیمه",
        "امتیاز رضایت",
        "آخرین فعالیت",
        "تاریخ ایجاد",
    ]
    worksheet.append(headers)
    _style_header(worksheet)
    total_messages = 0
    breached = 0
    for ticket in tickets:
        total_messages += ticket.message_count
        breached += 1 if ticket.sla_breached_at else 0
        worksheet.append(
            [
                ticket.ticket_number,
                ticket.subject,
                _display_user(ticket.owner),
                ticket.department.title,
                ticket.category.title,
                ticket.ticket_type.title,
                ticket.get_status_display(),
                ticket.get_priority_display(),
                ticket.get_severity_display(),
                _display_user(ticket.assigned_to) if ticket.assigned_to else "",
                "بله" if ticket.sla_breached_at else "خیر",
                ticket.message_count,
                ticket.attachment_count,
                ticket.satisfaction_rating_snapshot or "",
                _format_dt(ticket.last_activity_at),
                _format_dt(ticket.created_at),
            ]
        )
    worksheet.append(["جمع", "", "", "", "", "", "", "", "", "", breached, total_messages, "", "", "", ""])
    _style_body(worksheet, include_summary=True)
    _set_widths(worksheet, [20, 36, 28, 24, 24, 22, 18, 16, 16, 28, 16, 14, 14, 14, 24, 24])
    return _save(workbook)


def build_messages_workbook(*, messages: Iterable[SupportTicketMessage]) -> BytesIO:
    """Build an RTL Excel workbook for ticket timeline messages."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "پیام‌ها"
    _prepare_sheet(worksheet)
    headers = ["شماره تیکت", "نوع پیام", "نویسنده", "داخلی؟", "متن", "زمان ایجاد"]
    worksheet.append(headers)
    _style_header(worksheet)
    for message in messages:
        worksheet.append(
            [
                message.ticket.ticket_number,
                message.get_message_type_display(),
                _display_user(message.author),
                "بله" if message.is_internal else "خیر",
                message.body,
                _format_dt(message.created_at),
            ]
        )
    _style_body(worksheet)
    _set_widths(worksheet, [20, 22, 28, 12, 70, 24])
    return _save(workbook)


def build_sla_workbook(*, tickets: Iterable[SupportTicket]) -> BytesIO:
    """Build an RTL Excel workbook for SLA breaches and deadlines."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "SLA"
    _prepare_sheet(worksheet)
    headers = ["شماره تیکت", "وضعیت", "سیاست SLA", "اولین پاسخ تا", "حل تا", "نقض در", "ثانیه توقف", "ارجاع فوری"]
    worksheet.append(headers)
    _style_header(worksheet)
    for ticket in tickets:
        worksheet.append(
            [
                ticket.ticket_number,
                ticket.get_status_display(),
                ticket.applied_sla_policy.title if ticket.applied_sla_policy else "",
                _format_dt(ticket.first_response_due_at),
                _format_dt(ticket.resolution_due_at),
                _format_dt(ticket.sla_breached_at),
                ticket.sla_total_paused_seconds,
                "بله" if ticket.escalated_at else "خیر",
            ]
        )
    _style_body(worksheet)
    _set_widths(worksheet, [20, 18, 28, 24, 24, 24, 18, 16])
    return _save(workbook)


def build_csat_workbook(*, ratings: Iterable[SupportTicketSatisfaction]) -> BytesIO:
    """Build an RTL Excel workbook for CSAT ratings."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "رضایت‌سنجی"
    _prepare_sheet(worksheet)
    headers = ["شماره تیکت", "کاربر", "امتیاز", "نظر", "زمان ثبت"]
    worksheet.append(headers)
    _style_header(worksheet)
    total = 0
    count = 0
    for rating in ratings:
        total += rating.rating
        count += 1
        worksheet.append([rating.ticket.ticket_number, _display_user(rating.user), rating.rating, rating.comment, _format_dt(rating.created_at)])
    worksheet.append(["میانگین", "", round(total / count, 2) if count else 0, "", ""])
    _style_body(worksheet, include_summary=True)
    _set_widths(worksheet, [20, 28, 12, 60, 24])
    return _save(workbook)


def build_support_export_filename(*, export_type: str) -> str:
    """Build deterministic support export filename."""
    return f"support-{export_type}-{timezone.now():%Y%m%d}.xlsx"


def _prepare_sheet(worksheet) -> None:
    """Apply common worksheet settings."""
    worksheet.sheet_view.rightToLeft = True
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = "A1:Z1"


def _style_header(worksheet) -> None:
    """Style header row."""
    for cell in worksheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _style_body(worksheet, *, include_summary: bool = False) -> None:
    """Style body rows and optional summary row."""
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = _BODY_FONT
            cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
    if include_summary and worksheet.max_row >= 2:
        for cell in worksheet[worksheet.max_row]:
            cell.fill = _SUMMARY_FILL
            cell.font = Font(name="Tahoma", bold=True)


def _set_widths(worksheet, widths: list[int]) -> None:
    """Set worksheet column widths."""
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width


def _save(workbook: Workbook) -> BytesIO:
    """Save workbook to BytesIO."""
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _format_dt(value) -> str:
    """Format datetime for spreadsheet cells."""
    if value is None:
        return ""
    return timezone.localtime(value).strftime("%Y-%m-%d %H:%M")


def _display_user(user) -> str:
    """Return safe user display string."""
    if user is None:
        return ""
    return getattr(user, "full_name", "") or getattr(user, "email", "") or str(user)
