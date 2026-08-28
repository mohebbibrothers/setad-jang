"""Kindness Wall Phase 4 admin, export, and analytics tests."""

from __future__ import annotations

from io import BytesIO
from unittest.mock import MagicMock, patch

import pytest
from django.urls import reverse
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.test import APIClient

from apps.audit_logs import actions as audit_actions
from apps.kindness_wall.choices import DuplicateStatus, ListingStatus
from apps.kindness_wall.models import (
    KindnessContactReveal,
    KindnessDuplicateCandidate,
    KindnessListingReport,
    KindnessMatch,
)
from tests.factories import AdminUserFactory
from tests.factories.kindness_wall import (
    KindnessCategoryFactory,
    KindnessUserFactory,
    PublishedNeedListingFactory,
    PublishedOfferListingFactory,
)

pytestmark = pytest.mark.django_db

_AUDIT_TASK_PATH = "apps.audit_logs.tasks.create_audit_log_task"


def _client_for(user) -> APIClient:
    """Return authenticated API client."""
    client = APIClient()
    client.force_authenticate(user=user)
    return client


class TestKindnessAdminCategoryManagement:
    """Admin CRUD for professional tree categories."""

    def test_admin_can_create_move_and_soft_delete_category_with_audit(self) -> None:
        admin = AdminUserFactory()
        root = KindnessCategoryFactory(title="ریشه مهربانی")
        client = _client_for(admin)

        with patch(_AUDIT_TASK_PATH) as mock_task:
            mock_task.delay = MagicMock()
            create_response = client.post(
                reverse("kindness_wall:admin-category-list-create"),
                data={
                    "parent_id": root.pk,
                    "title": "زیرشاخه حرفه‌ای",
                    "description": "برای تست",
                    "icon": "briefcase",
                    "order": 5,
                },
                format="json",
            )

        assert create_response.status_code == status.HTTP_201_CREATED
        child_id = create_response.data["data"]["id"]
        assert create_response.data["data"]["parent_id"] == root.pk
        assert mock_task.delay.call_args.kwargs["action"] == audit_actions.KINDNESS_CATEGORY_CREATED

        update_response = client.patch(
            reverse("kindness_wall:admin-category-detail", kwargs={"category_id": child_id}),
            data={"parent_id": None, "order": 1},
            format="json",
        )
        assert update_response.status_code == status.HTTP_200_OK
        assert update_response.data["data"]["parent_id"] is None
        assert update_response.data["data"]["depth"] == 0

        delete_response = client.delete(
            reverse("kindness_wall:admin-category-detail", kwargs={"category_id": child_id})
        )
        assert delete_response.status_code == status.HTTP_200_OK

    def test_admin_cannot_create_category_cycle(self) -> None:
        admin = AdminUserFactory()
        root = KindnessCategoryFactory(title="والد")
        child = KindnessCategoryFactory(title="فرزند", parent=root)

        response = _client_for(admin).patch(
            reverse("kindness_wall:admin-category-detail", kwargs={"category_id": root.pk}),
            data={"parent_id": child.pk},
            format="json",
        )

        assert response.status_code == status.HTTP_400_BAD_REQUEST


class TestKindnessAdminOperationalQueues:
    """Admin match/contact/duplicate queues."""

    def test_admin_can_filter_matches_and_retrieve_detail(self) -> None:
        category = KindnessCategoryFactory(title="مشاغل")
        source = PublishedNeedListingFactory(category=category)
        target = PublishedOfferListingFactory(category=category)
        match = KindnessMatch.objects.create(source_listing=source, target_listing=target, score=86)
        client = _client_for(AdminUserFactory())

        list_response = client.get(
            reverse("kindness_wall:admin-match-list"),
            data={"min_score": 80, "category": category.slug},
        )
        detail_response = client.get(
            reverse("kindness_wall:admin-match-detail", kwargs={"match_id": match.pk})
        )

        assert list_response.status_code == status.HTTP_200_OK
        assert list_response.data["data"]["count"] == 1
        assert detail_response.status_code == status.HTTP_200_OK
        assert detail_response.data["data"]["id"] == match.pk

    def test_admin_can_list_contact_reveals_without_user_agent_bulk_noise(self) -> None:
        listing = PublishedNeedListingFactory(contact_phone_snapshot="+989120000000")
        viewer = KindnessUserFactory()
        KindnessContactReveal.objects.create(
            listing=listing,
            viewer=viewer,
            listing_owner=listing.owner,
            phone_snapshot=listing.contact_phone_snapshot,
            ip_address="127.0.0.1",
            user_agent="test-agent",
        )

        response = _client_for(AdminUserFactory()).get(
            reverse("kindness_wall:admin-contact-reveal-list"), data={"listing_id": listing.pk}
        )

        assert response.status_code == status.HTTP_200_OK
        row = response.data["data"]["results"][0]
        assert row["listing_id"] == listing.pk
        assert row["phone_snapshot"] == "+989120000000"
        assert "user_agent" not in row

    def test_admin_can_review_duplicate_candidate_with_audit(self) -> None:
        listing = PublishedNeedListingFactory(title="نیاز به برنامه نویس Django")
        candidate = PublishedNeedListingFactory(
            owner=listing.owner, category=listing.category, title="نیاز به برنامه نویس جنگو"
        )
        duplicate = KindnessDuplicateCandidate.objects.create(
            listing=listing, candidate_listing=candidate, score=91
        )

        with patch(_AUDIT_TASK_PATH) as mock_task:
            mock_task.delay = MagicMock()
            response = _client_for(AdminUserFactory()).post(
                reverse(
                    "kindness_wall:admin-duplicate-review", kwargs={"duplicate_id": duplicate.pk}
                ),
                data={"status": DuplicateStatus.CONFIRMED, "reason": "متن و مالک یکسان است"},
                format="json",
            )

        assert response.status_code == status.HTTP_200_OK
        duplicate.refresh_from_db()
        assert duplicate.status == DuplicateStatus.CONFIRMED
        assert (
            mock_task.delay.call_args.kwargs["action"] == audit_actions.KINDNESS_DUPLICATE_REVIEWED
        )


class TestKindnessAdminAnalyticsAndExports:
    """Advanced analytics and Excel exports."""

    def test_analytics_contains_distributions_and_match_effectiveness(self) -> None:
        category = KindnessCategoryFactory(title="سلامت")
        source = PublishedNeedListingFactory(category=category, province="تهران", city="ری")
        target = PublishedOfferListingFactory(category=category, province="تهران", city="ری")
        KindnessMatch.objects.create(
            source_listing=source, target_listing=target, score=78, status="contacted"
        )

        response = _client_for(AdminUserFactory()).get(reverse("kindness_wall:admin-analytics"))

        assert response.status_code == status.HTTP_200_OK
        data = response.data["data"]
        assert data["total_listings"] >= 2
        assert data["match_effectiveness"]["contacted_matches"] == 1
        assert data["province_distribution"][0]["province"] == "تهران"
        assert data["category_distribution"][0]["category__title"] == "سلامت"

    def test_admin_listing_export_returns_rtl_xlsx_and_audit(self) -> None:
        listing = PublishedNeedListingFactory(
            title="خروجی اکسل دیوار", view_count=12, contact_reveal_count=3
        )

        with patch(_AUDIT_TASK_PATH) as mock_task:
            mock_task.delay = MagicMock()
            response = _client_for(AdminUserFactory()).get(
                reverse("kindness_wall:admin-listing-export"),
                data={"status": ListingStatus.PUBLISHED},
            )

        assert response.status_code == status.HTTP_200_OK
        assert (
            response["Content-Type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook.active
        assert sheet.sheet_view.rightToLeft is True
        assert sheet["D2"].value == listing.title
        assert sheet["A3"].value == "جمع"
        assert (
            mock_task.delay.call_args.kwargs["action"] == audit_actions.KINDNESS_LISTINGS_EXPORTED
        )

    def test_admin_report_export_returns_rtl_xlsx_and_audit(self) -> None:
        listing = PublishedNeedListingFactory(title="آگهی گزارش‌دار")
        report = KindnessListingReport.objects.create(
            listing=listing,
            reported_by=KindnessUserFactory(),
            reason="other",
            description="بررسی شود",
        )

        with patch(_AUDIT_TASK_PATH) as mock_task:
            mock_task.delay = MagicMock()
            response = _client_for(AdminUserFactory()).get(
                reverse("kindness_wall:admin-report-export"), data={"status": report.status}
            )

        assert response.status_code == status.HTTP_200_OK
        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook.active
        assert sheet.sheet_view.rightToLeft is True
        assert sheet["C2"].value == "آگهی گزارش‌دار"
        assert mock_task.delay.call_args.kwargs["action"] == audit_actions.KINDNESS_REPORTS_EXPORTED
