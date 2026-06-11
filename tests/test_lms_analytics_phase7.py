"""
LMS Phase 7 analytics/reporting/export tests.

این تست‌ها گزارش ادمین کلاس را از سطح basic report به analytics حرفه‌ای‌تر ارتقا
می‌دهند: summary، leaderboard، مدال‌ها و export اکسل قابل دانلود.
"""

from __future__ import annotations

from decimal import Decimal
from io import BytesIO
from unittest.mock import MagicMock, patch

import pytest
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.test import APIClient

from apps.audit_logs import actions as audit_actions
from apps.lms.choices import (
    BadgeLevel,
    CertificateStatus,
    EnrollmentStatus,
    QuizAttemptStatus,
    QuizStatus,
)
from apps.lms.models import Certificate, Enrollment, LMSUserSkill, QuizAttempt
from tests.factories import AdminUserFactory, UserFactory
from tests.factories.lms import PublishedCourseFactory, QuizFactory

pytestmark = pytest.mark.django_db

_AUDIT_TASK_PATH = "apps.audit_logs.tasks.create_audit_log_task"


def _client_for(user) -> APIClient:
    """Return authenticated API client."""
    client = APIClient()
    client.force_authenticate(user=user)
    return client


def _make_user(email: str, *, first_name: str = "Ali", last_name: str = "User"):
    """Create user with profile data useful for LMS reports."""
    user = UserFactory(email=email, first_name=first_name, last_name=last_name)
    user.profile.national_code = "0123456789"
    user.profile.save(update_fields=["national_code"])
    return user


def _make_enrollment(*, course, user, status_value=EnrollmentStatus.ACTIVE, progress=Decimal("0.00")):
    """Create enrollment row for analytics tests."""
    return Enrollment.objects.create(
        course=course,
        user=user,
        status=status_value,
        progress_percent=progress,
        watched_seconds=int(progress),
        total_seconds_snapshot=100,
    )


def _make_attempt(*, course, enrollment, user, score: Decimal, passed: bool):
    """Create a submitted quiz attempt for analytics tests."""
    quiz, _ = QuizFactory._get_manager(QuizFactory._meta.model).get_or_create(
        course=course,
        defaults={"title": "آزمون", "status": QuizStatus.PUBLISHED},
    )
    return QuizAttempt.objects.create(
        quiz=quiz,
        course=course,
        enrollment=enrollment,
        user=user,
        attempt_number=1,
        status=QuizAttemptStatus.PASSED if passed else QuizAttemptStatus.FAILED,
        score_out_of_20=score,
        score_percent=score * Decimal("5"),
        score_raw=score,
        is_passed=passed,
        submitted_at=timezone.now(),
    )


def _issue_certificate(*, course, enrollment, attempt, user, badge=BadgeLevel.GOLD):
    """Create certificate and skill rows for leaderboard tests."""
    certificate = Certificate.objects.create(
        user=user,
        course=course,
        enrollment=enrollment,
        quiz_attempt=attempt,
        status=CertificateStatus.ISSUED,
        full_name_snapshot=user.full_name,
        gender_snapshot="male",
        national_code_snapshot="0123456789",
        course_title_snapshot=course.title,
        instructor_name_snapshot=course.instructor_name,
        score_out_of_20=attempt.score_out_of_20,
    )
    LMSUserSkill.objects.create(
        user=user,
        course=course,
        certificate=certificate,
        title=course.title,
        badge_level=badge,
    )
    return certificate


class TestLMSCourseAnalytics:
    """Admin course analytics endpoint tests."""

    def test_admin_course_analytics_returns_counts_and_average_score(self) -> None:
        admin = AdminUserFactory()
        course = PublishedCourseFactory(title="تحلیل کلاس")
        user_a = _make_user("a@example.com", first_name="A")
        user_b = _make_user("b@example.com", first_name="B")
        enrollment_a = _make_enrollment(course=course, user=user_a, status_value=EnrollmentStatus.COMPLETED, progress=Decimal("100.00"))
        enrollment_b = _make_enrollment(course=course, user=user_b, status_value=EnrollmentStatus.ACTIVE, progress=Decimal("50.00"))
        attempt_a = _make_attempt(course=course, enrollment=enrollment_a, user=user_a, score=Decimal("18.00"), passed=True)
        _make_attempt(course=course, enrollment=enrollment_b, user=user_b, score=Decimal("10.00"), passed=False)
        _issue_certificate(course=course, enrollment=enrollment_a, attempt=attempt_a, user=user_a)

        response = _client_for(admin).get(reverse("lms:admin-course-analytics", kwargs={"course_id": course.pk}))

        assert response.status_code == status.HTTP_200_OK
        data = response.data["data"]
        assert data["participants_count"] == 2
        assert data["graduates_count"] == 1
        assert data["quiz_attempts_count"] == 2
        assert data["quiz_passed_count"] == 1
        assert data["quiz_failed_count"] == 1
        assert data["average_score_out_of_20"] == 14.0

    def test_admin_course_leaderboard_orders_by_score_and_badge(self) -> None:
        admin = AdminUserFactory()
        course = PublishedCourseFactory(title="رتبه‌بندی")
        winner = _make_user("winner@example.com", first_name="Winner")
        runner = _make_user("runner@example.com", first_name="Runner")
        winner_enrollment = _make_enrollment(course=course, user=winner, status_value=EnrollmentStatus.COMPLETED, progress=Decimal("100.00"))
        runner_enrollment = _make_enrollment(course=course, user=runner, status_value=EnrollmentStatus.COMPLETED, progress=Decimal("100.00"))
        winner_attempt = _make_attempt(course=course, enrollment=winner_enrollment, user=winner, score=Decimal("19.50"), passed=True)
        runner_attempt = _make_attempt(course=course, enrollment=runner_enrollment, user=runner, score=Decimal("17.00"), passed=True)
        _issue_certificate(course=course, enrollment=winner_enrollment, attempt=winner_attempt, user=winner, badge=BadgeLevel.DISTINCTION)
        _issue_certificate(course=course, enrollment=runner_enrollment, attempt=runner_attempt, user=runner, badge=BadgeLevel.SILVER)

        response = _client_for(admin).get(reverse("lms:admin-course-leaderboard", kwargs={"course_id": course.pk}))

        assert response.status_code == status.HTTP_200_OK
        rows = response.data["data"]
        assert rows[0]["user_id"] == winner.pk
        assert rows[0]["badge_level"] == BadgeLevel.DISTINCTION
        assert rows[0]["best_score_out_of_20"] == 19.5


class TestLMSCourseExport:
    """Admin Excel export tests."""

    def test_admin_course_export_returns_xlsx_and_dispatches_audit(self) -> None:
        admin = AdminUserFactory()
        course = PublishedCourseFactory(title="خروجی اکسل")
        user = _make_user("export@example.com", first_name="Export", last_name="User")
        _make_enrollment(course=course, user=user, status_value=EnrollmentStatus.ACTIVE, progress=Decimal("42.00"))

        with patch(_AUDIT_TASK_PATH) as mock_task:
            mock_task.delay = MagicMock()
            response = _client_for(admin).get(reverse("lms:admin-course-export", kwargs={"course_id": course.pk}))

        assert response.status_code == status.HTTP_200_OK
        assert response["Content-Type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert "attachment" in response["Content-Disposition"]
        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook.active
        assert sheet.sheet_view.rightToLeft is True
        assert sheet.max_row == 2
        assert sheet["C2"].value == "export@example.com"
        assert mock_task.delay.call_args.kwargs["action"] == audit_actions.LMS_COURSE_REPORT_EXPORTED
