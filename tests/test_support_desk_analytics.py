"""Support Desk Phase 5 analytics, export and SLA task tests."""

from __future__ import annotations

from io import BytesIO
from unittest.mock import MagicMock, patch

import pytest
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.test import APIClient

from apps.audit_logs import actions as audit_actions
from apps.support_desk.choices import TicketStatus
from apps.support_desk.models import SupportSLAEvent, SupportTicket, SupportTicketSatisfaction
from apps.support_desk.services import (
    add_admin_reply,
    create_ticket,
    resolve_ticket,
    submit_ticket,
)
from apps.support_desk.tasks import (
    cleanup_stale_support_drafts_task,
    daily_support_digest_task,
    mark_support_sla_breaches_task,
)
from tests.factories import AdminUserFactory, UserFactory
from tests.factories.support_desk import (
    SupportSLAPolicyFactory,
    SupportTicketFactory,
    SupportTicketTypeFactory,
)

pytestmark = pytest.mark.django_db

_AUDIT_TASK_PATH = "apps.audit_logs.tasks.create_audit_log_task"


def _client_for(user) -> APIClient:
    """Return authenticated API client."""
    client = APIClient()
    client.force_authenticate(user=user)
    return client


def _make_sla_ticket(*, breached: bool = False) -> SupportTicket:
    """Create a submitted ticket with an applied SLA policy."""
    policy = SupportSLAPolicyFactory(
        title="SLA گزارش", first_response_minutes=1, resolution_minutes=2
    )
    ticket_type = SupportTicketTypeFactory(default_sla_policy=policy)
    ticket = create_ticket(
        owner=UserFactory(),
        ticket_type=ticket_type,
        subject="گزارش پرداخت",
        description="پرداخت انجام شد اما ثبت نشد",
    )
    now = timezone.now() - timezone.timedelta(minutes=10) if breached else timezone.now()
    submit_ticket(ticket=ticket, user=ticket.owner, now=now)
    return ticket


class TestSupportAdminAnalytics:
    """Admin analytics dashboard tests."""

    def test_admin_analytics_contains_rates_distributions_and_csat(self) -> None:
        ticket = _make_sla_ticket()
        admin = AdminUserFactory()
        resolve_ticket(ticket=ticket, admin=admin)
        SupportTicketSatisfaction.objects.create(
            ticket=ticket, user=ticket.owner, rating=5, comment="عالی"
        )
        ticket.satisfaction_rating_snapshot = 5
        ticket.save(update_fields=["satisfaction_rating_snapshot", "updated_at"])

        response = _client_for(admin).get(reverse("support_desk:admin-analytics"))

        assert response.status_code == status.HTTP_200_OK
        data = response.data["data"]
        assert data["total_tickets"] >= 1
        assert data["resolved_tickets"] >= 1
        assert data["csat_average"] == 5
        assert data["status_distribution"]
        assert data["department_distribution"]
        assert data["generated_at"]


class TestSupportAdminExports:
    """Excel export endpoint tests."""

    def test_ticket_export_returns_rtl_xlsx_and_dispatches_audit(self) -> None:
        ticket = _make_sla_ticket()
        admin = AdminUserFactory()

        with patch(_AUDIT_TASK_PATH) as mock_task:
            mock_task.delay = MagicMock()
            response = _client_for(admin).get(
                reverse("support_desk:admin-export-tickets"), data={"search": ticket.ticket_number}
            )

        assert response.status_code == status.HTTP_200_OK
        assert (
            response["Content-Type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook.active
        assert sheet.sheet_view.rightToLeft is True
        assert sheet["A2"].value == ticket.ticket_number
        assert mock_task.delay.call_args.kwargs["action"] == audit_actions.SUPPORT_EXPORT_GENERATED

    def test_messages_sla_and_csat_exports_return_rtl_xlsx(self) -> None:
        ticket = _make_sla_ticket()
        admin = AdminUserFactory()
        add_admin_reply(ticket=ticket, admin=admin, body="پاسخ گزارش")
        resolve_ticket(ticket=ticket, admin=admin)
        SupportTicketSatisfaction.objects.create(ticket=ticket, user=ticket.owner, rating=4)

        client = _client_for(admin)
        for route_name in ["admin-export-messages", "admin-export-sla", "admin-export-csat"]:
            response = client.get(reverse(f"support_desk:{route_name}"))
            assert response.status_code == status.HTTP_200_OK
            workbook = load_workbook(BytesIO(response.content))
            assert workbook.active.sheet_view.rightToLeft is True


class TestSupportTasks:
    """Celery task behavior tests."""

    def test_mark_sla_breaches_task_marks_and_escalates_due_ticket(self) -> None:
        ticket = _make_sla_ticket(breached=True)

        updated = mark_support_sla_breaches_task()
        ticket.refresh_from_db()

        assert updated >= 1
        assert ticket.sla_breached_at is not None
        assert ticket.status == TicketStatus.ESCALATED
        assert SupportSLAEvent.objects.filter(ticket=ticket).exists()

    def test_cleanup_stale_support_drafts_task_archives_old_drafts(self) -> None:
        old_ticket = SupportTicketFactory(status=TicketStatus.DRAFT)
        SupportTicket.objects.filter(pk=old_ticket.pk).update(
            created_at=timezone.now() - timezone.timedelta(days=45)
        )

        updated = cleanup_stale_support_drafts_task(older_than_days=30)
        old_ticket.refresh_from_db()

        assert updated == 1
        assert old_ticket.status == TicketStatus.ARCHIVED

    def test_daily_support_digest_task_returns_monitoring_counters(self) -> None:
        SupportTicketFactory(status=TicketStatus.SUBMITTED, assigned_to=None)

        digest = daily_support_digest_task()

        assert digest["open_tickets"] >= 1
        assert digest["unassigned_tickets"] >= 1
        assert "sla_breached_tickets" in digest
