"""Regression tests for the high-priority (P1) hardening work.

هر تست اینجا به یک یافتهٔ مشخص از گزارش آنالیز گره خورده و طوری نوشته شده
که اگر آن اصلاح برگردد قرمز شود. تست‌ها تا حد امکان *رفتاری* هستند: به‌جای
grep روی سورس، یا خروجی واقعی را می‌سنجند یا با AST ساختار کد را بررسی
می‌کنند.
"""

from __future__ import annotations

import ast
import inspect
import subprocess
import sys
from pathlib import Path

import pytest
import yaml
from django.conf import settings
from django.core.cache import cache
from django.db import connection
from django.test import RequestFactory
from django.test.utils import CaptureQueriesContext
from django.urls import reverse
from django.utils import timezone
from rest_framework.parsers import JSONParser
from rest_framework.request import Request
from rest_framework.test import APIClient

from tests.factories import AdminUserFactory, UserFactory

REPO_ROOT = Path(__file__).resolve().parent.parent


def _decorator_names(module, function_name: str) -> list[str]:
    """Return decorator names on a top-level function, via AST."""
    tree = ast.parse(inspect.getsource(module))
    node = next(
        item for item in tree.body
        if isinstance(item, (ast.FunctionDef, ast.AsyncFunctionDef)) and item.name == function_name
    )
    names = []
    for decorator in node.decorator_list:
        target = decorator.func if isinstance(decorator, ast.Call) else decorator
        names.append(target.attr if isinstance(target, ast.Attribute) else getattr(target, "id", "?"))
    return names


# ===========================================================================
# 4.1 — cache key flooding via arbitrary query params
# ===========================================================================

class TestCacheKeyBounding:
    """The public cache key must be bounded and canonical."""

    @staticmethod
    def _variant(query: str):
        """Build the cache-key variant tuple for one query string."""
        from apps.core.api_cache import build_cache_variant
        from apps.core.pagination import StandardPagination
        from apps.madadkar.filters import CampaignPublicFilter

        request = Request(RequestFactory().get(f"/x?{query}"), parsers=[JSONParser()])
        filterset = CampaignPublicFilter(request.query_params, queryset=None)
        return build_cache_variant(request, filterset=filterset, pagination_class=StandardPagination)

    def test_unknown_query_params_do_not_create_new_cache_keys(self) -> None:
        """The classic flooding attack must collapse to a single key."""
        variants = {self._variant(f"zzz={index}") for index in range(500)}
        assert variants == {self._variant("")}

    def test_oversized_page_size_collapses_to_the_clamped_value(self) -> None:
        """page_size above max_page_size yields one response, so one key."""
        from apps.core.pagination import StandardPagination

        variants = {self._variant(f"page_size={size}") for size in range(101, 400)}
        assert len(variants) == 1
        assert variants == {self._variant(f"page_size={StandardPagination.max_page_size}")}

    def test_page_number_spellings_are_canonicalized(self) -> None:
        """Different spellings of page 1 must not each cost a cache miss."""
        variants = {
            self._variant("page=1"),
            self._variant("page=001"),
            self._variant("page=+1"),
            self._variant("page=%201%20"),
            self._variant("page="),
            self._variant(""),
        }
        assert len(variants) == 1

    def test_invalid_filter_values_collapse_to_the_unfiltered_key(self) -> None:
        """An invalid filterset falls back to the unfiltered queryset in the view."""
        variants = {self._variant(f"status=bogus{index}") for index in range(200)}
        assert variants == {self._variant("")}

    def test_genuinely_different_filters_stay_distinct(self) -> None:
        """Bounding must never merge requests that produce different data."""
        published = self._variant("status=published")
        completed = self._variant("status=completed")
        unfiltered = self._variant("")
        assert len({published, completed, unfiltered}) == 3

    def test_distinct_pages_stay_distinct(self) -> None:
        """Canonicalization must never merge two different pages."""
        assert len({self._variant("page=1"), self._variant("page=2"), self._variant("page=3")}) == 3

    def test_last_page_keyword_keeps_its_own_identity(self) -> None:
        """DRF treats page=last specially, so it must not fall into the invalid bucket."""
        from apps.core.api_cache import canonical_page
        from apps.core.pagination import StandardPagination

        request = Request(RequestFactory().get("/x?page=last"), parsers=[JSONParser()])
        assert canonical_page(request, pagination_class=StandardPagination) == "last"
        assert self._variant("page=last") != self._variant("page=abc")

    def test_empty_filter_value_is_treated_as_no_filter(self) -> None:
        """django-filter ignores empty values, so the cache key must too."""
        assert self._variant("status=") == self._variant("")


# ===========================================================================
# 4.2 — production must refuse to boot on a per-process cache
# ===========================================================================

class TestProductionCacheFailFast:
    """Production settings must reject non-shared cache backends."""

    @staticmethod
    def _boot(cache_backend: str) -> subprocess.CompletedProcess:
        """Import production settings in a subprocess with a given CACHE_BACKEND."""
        env = {
            "PATH": "/usr/bin:/bin",
            "DJANGO_SETTINGS_MODULE": "config.settings.production",
            "ALLOWED_HOSTS": "example.com",
            "CACHE_BACKEND": cache_backend,
            "CORS_ALLOWED_ORIGINS": "https://example.com",
            "DATABASE_ENGINE": "postgres",
            "POSTGRES_DB": "x",
            "POSTGRES_USER": "x",
            "POSTGRES_PASSWORD": "strong-postgres-password",
            "POSTGRES_HOST": "127.0.0.1",
            "POSTGRES_PORT": "5432",
            "SECRET_KEY": "realistic-production-secret-key-with-more-than-fifty-characters-2026",
        }
        return subprocess.run(
            [sys.executable, "-c", "import django; django.setup()"],
            cwd=REPO_ROOT,
            env=env,
            capture_output=True,
            text=True,
            timeout=120,
            check=False,
        )

    def test_locmem_is_rejected_in_production(self) -> None:
        """A per-process cache silently weakens every throttle, so it must fail fast."""
        result = self._boot("locmem")
        assert result.returncode != 0
        assert "CACHE_BACKEND" in result.stderr

    def test_redis_is_accepted_in_production(self) -> None:
        """The guard must not block a correct configuration."""
        assert self._boot("redis").returncode == 0


# ===========================================================================
# 4.3 — gunicorn worker model
# ===========================================================================

class TestGunicornConfiguration:
    """The container command must survive blocking I/O and deploys."""

    @staticmethod
    def _dockerfile() -> str:
        """Return the Dockerfile contents."""
        return (REPO_ROOT / "Dockerfile").read_text(encoding="utf-8")

    def test_uses_threaded_workers(self) -> None:
        """Sync workers cap the whole site at --workers concurrent requests."""
        content = self._dockerfile()
        assert "--worker-class ${GUNICORN_WORKER_CLASS:-gthread}" in content
        assert "--threads ${GUNICORN_THREADS:-8}" in content

    def test_execs_gunicorn_so_sigterm_is_delivered(self) -> None:
        """Without exec, the shell swallows SIGTERM and deploys kill live requests."""
        assert "exec gunicorn" in self._dockerfile()

    def test_worker_recycling_and_graceful_shutdown_are_configured(self) -> None:
        """Guards against slow memory leaks and abrupt shutdowns."""
        content = self._dockerfile()
        assert "--max-requests" in content
        assert "--max-requests-jitter" in content
        assert "--graceful-timeout" in content
        assert "--worker-tmp-dir /dev/shm" in content

    def test_access_log_carries_the_request_id(self) -> None:
        """Correlating gunicorn logs with app logs requires the middleware request id."""
        assert "%({x-request-id}o)s" in self._dockerfile()


# ===========================================================================
# 4.4 — notification fan-out
# ===========================================================================

@pytest.mark.django_db
class TestNotificationFanOut:
    """Notification creation must not scale queries with recipient count."""

    def test_query_count_is_independent_of_recipient_count(self) -> None:
        """The whole point of the fix: constant queries, not 5 x N x C."""
        from apps.notifications.choices import NotificationChannel
        from apps.notifications.services import create_notification_event

        channels = [NotificationChannel.IN_APP, NotificationChannel.EMAIL]
        small = [UserFactory() for _ in range(5)]
        large = [UserFactory() for _ in range(40)]

        with CaptureQueriesContext(connection) as small_ctx:
            create_notification_event(event_type="fanout.small", recipients=small, channels=channels, payload={"title": "t"})
        with CaptureQueriesContext(connection) as large_ctx:
            create_notification_event(event_type="fanout.large", recipients=large, channels=channels, payload={"title": "t"})

        assert len(small_ctx.captured_queries) == len(large_ctx.captured_queries)
        assert len(large_ctx.captured_queries) <= 12

    def test_disabled_preferences_are_still_respected(self) -> None:
        """Batching must not change who receives a notification."""
        from apps.notifications.choices import NotificationChannel
        from apps.notifications.services import create_notification_event, set_preference

        allowed = UserFactory()
        muted = UserFactory()
        set_preference(user=muted, event_type="pref.check", channel=NotificationChannel.IN_APP, enabled=False)

        event = create_notification_event(
            event_type="pref.check",
            recipients=[allowed, muted],
            channels=[NotificationChannel.IN_APP],
            payload={"title": "t"},
        )
        assert [delivery.recipient_id for delivery in event.deliveries.all()] == [allowed.pk]

    def test_duplicate_recipients_do_not_raise(self) -> None:
        """bulk_create has no get_or_create safety net, so dedupe must be airtight."""
        from apps.notifications.choices import NotificationChannel
        from apps.notifications.services import create_notification_event

        user = UserFactory()
        event = create_notification_event(
            event_type="dupe.check",
            recipients=[user, user, user],
            channels=[NotificationChannel.IN_APP, NotificationChannel.IN_APP],
            payload={"title": "t"},
        )
        assert event.deliveries.count() == 1

    def test_activity_rows_are_still_created(self) -> None:
        """Moving to bulk_create must not silently drop the activity timeline."""
        from apps.activity.models import UserActivity
        from apps.notifications.choices import NotificationChannel
        from apps.notifications.services import create_notification_event

        user = UserFactory()
        create_notification_event(
            event_type="activity.check",
            recipients=[user],
            channels=[NotificationChannel.IN_APP],
            payload={"title": "عنوان"},
            aggregate_type="thing",
            aggregate_id="7",
        )
        activity = UserActivity.objects.get(user=user, event_type="activity.check")
        assert activity.title == "عنوان"
        assert activity.aggregate_id == "7"

    def test_dispatch_does_not_wrap_provider_io_in_a_transaction(self) -> None:
        """A rollback cannot un-send an email, so provider I/O must sit outside atomic."""
        from apps.notifications import services

        assert "atomic" not in _decorator_names(services, "dispatch_event")

    def test_unexpected_provider_error_fails_one_delivery_not_the_batch(self) -> None:
        """One broken channel must not lose the results of the healthy ones."""
        from unittest.mock import patch

        from apps.notifications.choices import NotificationChannel, NotificationDeliveryStatus
        from apps.notifications.services import create_notification_event, dispatch_event

        event = create_notification_event(
            event_type="boom.check",
            recipients=[UserFactory()],
            channels=[NotificationChannel.IN_APP],
            payload={"title": "t"},
        )
        with patch("apps.notifications.services.get_notification_provider", side_effect=RuntimeError("boom")):
            dispatch_event(event=event)

        delivery = event.deliveries.get()
        assert delivery.status == NotificationDeliveryStatus.FAILED
        assert delivery.error_message == "RuntimeError"

    def test_dispatch_does_not_refetch_recipients_per_delivery(self) -> None:
        """select_related must keep dispatch query count flat."""
        from apps.notifications.choices import NotificationChannel
        from apps.notifications.services import create_notification_event, dispatch_event

        event = create_notification_event(
            event_type="dispatch.n1",
            recipients=[UserFactory() for _ in range(10)],
            channels=[NotificationChannel.IN_APP],
            payload={"title": "t"},
        )
        with CaptureQueriesContext(connection) as ctx:
            dispatch_event(event=event)
        assert len(ctx.captured_queries) <= 8


# ===========================================================================
# 4.5 — command center
# ===========================================================================

@pytest.mark.django_db
class TestCommandCenter:
    """The admin dashboard must be cached and cheap."""

    def _get(self):
        """Call the command center endpoint as an admin."""
        client = APIClient()
        client.force_authenticate(user=AdminUserFactory())
        return client.get(reverse("command_center:summary"))

    def test_second_call_hits_the_cache(self) -> None:
        """Uncached COUNT(*) on multi-million-row tables cannot run per request."""
        cache.clear()
        with CaptureQueriesContext(connection) as cold_ctx:
            assert self._get().status_code == 200
        with CaptureQueriesContext(connection) as warm_ctx:
            assert self._get().status_code == 200
        assert len(warm_ctx.captured_queries) < len(cold_ctx.captured_queries)

    def test_counter_queries_are_collapsed_per_table(self) -> None:
        """Same-table counters must share one scan instead of one COUNT each."""
        from apps.command_center.selectors import build_command_center_counters

        cache.clear()
        with CaptureQueriesContext(connection) as ctx:
            build_command_center_counters()
        assert len(ctx.captured_queries) <= 30

    def test_health_section_is_never_served_from_cache(self) -> None:
        """An admin asking whether Redis is up needs a live answer."""
        from apps.command_center import selectors

        cache.clear()
        self._get()
        calls = []
        original = selectors._health_summary

        def spy():
            calls.append(1)
            return original()

        selectors._health_summary = spy
        try:
            self._get()
        finally:
            selectors._health_summary = original
        assert calls == [1]

    def test_response_exposes_counter_freshness(self) -> None:
        """Operators must be able to tell how stale the cached counters are."""
        cache.clear()
        data = self._get().data["data"]
        assert data["counters_generated_at"]
        assert data["generated_at"]


# ===========================================================================
# 4.6 — JWT blacklist growth
# ===========================================================================

@pytest.mark.django_db
class TestExpiredTokenFlush:
    """Rotated refresh tokens must be garbage collected."""

    def test_flush_task_is_scheduled(self) -> None:
        """Without a beat entry the tables grow forever."""
        tasks = {entry["task"] for entry in settings.CELERY_BEAT_SCHEDULE.values()}
        assert "apps.authentication.tasks.flush_expired_jwt_tokens_task" in tasks

    def test_only_expired_tokens_are_deleted(self) -> None:
        """A cleanup job that logs people out would be worse than the leak."""
        from rest_framework_simplejwt.token_blacklist.models import OutstandingToken

        from apps.authentication.tasks import flush_expired_jwt_tokens_task

        user = UserFactory()
        now = timezone.now()
        expired = OutstandingToken.objects.create(
            user=user, jti="expired-1", token="t1",
            created_at=now - timezone.timedelta(days=30),
            expires_at=now - timezone.timedelta(days=1),
        )
        alive = OutstandingToken.objects.create(
            user=user, jti="alive-1", token="t2",
            created_at=now, expires_at=now + timezone.timedelta(days=7),
        )

        result = flush_expired_jwt_tokens_task()

        assert result["deleted_outstanding"] == 1
        assert not OutstandingToken.objects.filter(pk=expired.pk).exists()
        assert OutstandingToken.objects.filter(pk=alive.pk).exists()

    def test_deletion_is_batched(self) -> None:
        """One unbounded DELETE on an accumulated table can lock out login."""
        from rest_framework_simplejwt.token_blacklist.models import OutstandingToken

        from apps.authentication.tasks import flush_expired_jwt_tokens_task

        user = UserFactory()
        now = timezone.now()
        OutstandingToken.objects.bulk_create([
            OutstandingToken(
                user=user, jti=f"e{index}", token=f"t{index}",
                created_at=now - timezone.timedelta(days=30),
                expires_at=now - timezone.timedelta(days=1),
            )
            for index in range(10)
        ])

        with CaptureQueriesContext(connection) as ctx:
            result = flush_expired_jwt_tokens_task(batch_size=2, max_batches=100)

        deletes = [q for q in ctx.captured_queries if q["sql"].lstrip().upper().startswith("DELETE")]
        assert result["deleted_outstanding"] == 10
        assert len(deletes) > 1

    def test_batch_cap_reports_unfinished_work(self) -> None:
        """A capped run must tell the operator that more rows remain."""
        from rest_framework_simplejwt.token_blacklist.models import OutstandingToken

        from apps.authentication.tasks import flush_expired_jwt_tokens_task

        user = UserFactory()
        now = timezone.now()
        OutstandingToken.objects.bulk_create([
            OutstandingToken(
                user=user, jti=f"c{index}", token=f"t{index}",
                created_at=now - timezone.timedelta(days=30),
                expires_at=now - timezone.timedelta(days=1),
            )
            for index in range(6)
        ])

        result = flush_expired_jwt_tokens_task(batch_size=2, max_batches=1)
        assert result["deleted_outstanding"] == 2
        assert result["exhausted"] == 0


# ===========================================================================
# 4.7 — Excel exports
# ===========================================================================

@pytest.mark.django_db
class TestStreamingExcelExports:
    """Exports must not materialize the whole workbook in a web worker."""

    def test_writer_is_write_only(self) -> None:
        """write_only mode is what keeps memory flat as rows grow."""
        from apps.core.excel import ExcelColumn, StreamingExcelSheet

        with StreamingExcelSheet(title="t", columns=[ExcelColumn("a")]) as sheet:
            assert sheet._workbook.write_only is True

    def test_row_cap_is_enforced(self) -> None:
        """A runaway export must fail loudly instead of eating the worker."""
        from apps.core.excel import ExcelColumn, ExcelExportTooLargeError, StreamingExcelSheet

        sheet = StreamingExcelSheet(title="t", columns=[ExcelColumn("a")], max_rows=3)
        for index in range(3):
            sheet.append([index])
        with pytest.raises(ExcelExportTooLargeError):
            sheet.append([99])
        # خطای سقف باید منبع را هم آزاد کرده باشد، نه فقط استثنا بدهد.
        assert sheet._saved is True

    def test_aware_datetimes_are_normalized(self) -> None:
        """openpyxl raises on tz-aware datetimes; the writer must absorb that."""
        from openpyxl import load_workbook

        from apps.core.excel import ExcelColumn, StreamingExcelSheet

        sheet = StreamingExcelSheet(title="t", columns=[ExcelColumn("when", kind="date")])
        sheet.append([timezone.now()])
        worksheet = load_workbook(sheet.save()).active
        assert worksheet.cell(row=2, column=1).value.tzinfo is None

    def test_stream_rows_uses_queryset_iterator(self) -> None:
        """Iterating a queryset directly fills _result_cache with every row."""
        from apps.core.excel import stream_rows
        from apps.madadkar.models import Campaign

        queryset = Campaign.objects.all()
        list(stream_rows(queryset, chunk_size=10))
        assert queryset._result_cache is None

    def test_stream_rows_accepts_plain_iterables(self) -> None:
        """Callers may still pass lists; the helper must not require a queryset."""
        from apps.core.excel import stream_rows

        assert list(stream_rows([1, 2, 3])) == [1, 2, 3]

    def test_campaign_export_streams_the_queryset(self) -> None:
        """The flagship export must not load all participants into memory."""
        from apps.madadkar import export

        source = inspect.getsource(export.generate_campaign_participants_excel)
        assert ".iterator(" in source
        assert "StreamingExcelSheet" in source

    def test_no_export_module_builds_a_full_workbook(self) -> None:
        """Every export module must go through the streaming writer."""
        offenders = []
        for name in ("kindness_wall", "lms", "madadkar", "support_desk"):
            source = (REPO_ROOT / "apps" / name / "export.py").read_text(encoding="utf-8")
            if "Workbook()" in source:
                offenders.append(name)
        assert offenders == []


# ===========================================================================
# 4.8 — Flower exposure
# ===========================================================================

class TestFlowerHardening:
    """The Celery dashboard leaks task arguments and can control workers."""

    @staticmethod
    def _flower_service() -> dict:
        """Return the raw flower service definition from docker-compose."""
        compose = yaml.safe_load((REPO_ROOT / "docker-compose.yml").read_text(encoding="utf-8"))
        return compose["services"]["flower"]

    def test_port_is_bound_to_loopback_only(self) -> None:
        """Publishing on 0.0.0.0 exposes payment authorities and audit payloads."""
        for mapping in self._flower_service()["ports"]:
            assert str(mapping).startswith("127.0.0.1:")

    def test_basic_auth_is_mandatory(self) -> None:
        """Compose must refuse to start without credentials."""
        command = " ".join(self._flower_service()["command"])
        assert "--basic-auth=" in command
        assert "FLOWER_USER:?" in command
        assert "FLOWER_PASSWORD:?" in command

    def test_healthcheck_authenticates(self) -> None:
        """An unauthenticated healthcheck would mark the container permanently unhealthy."""
        assert "-u" in " ".join(self._flower_service()["healthcheck"]["test"])


# ===========================================================================
# 4.9 — Celery delivery guarantees
# ===========================================================================

class TestCeleryDeliveryGuarantees:
    """Financial and audit tasks must not vanish when a worker dies."""

    def test_acks_late_is_enabled(self) -> None:
        """With early ack, a worker crash loses the task forever."""
        assert settings.CELERY_TASK_ACKS_LATE is True

    def test_reject_on_worker_lost_is_enabled(self) -> None:
        """Without it, acks_late does nothing in the exact crash case it targets."""
        assert settings.CELERY_TASK_REJECT_ON_WORKER_LOST is True

    def test_visibility_timeout_exceeds_the_task_time_limit(self) -> None:
        """Otherwise Redis redelivers a still-running task to a second worker."""
        visibility = settings.CELERY_BROKER_TRANSPORT_OPTIONS["visibility_timeout"]
        assert visibility > settings.CELERY_TASK_TIME_LIMIT


# ===========================================================================
# 4.10 — campaign counter recomputation
# ===========================================================================

@pytest.mark.django_db
class TestCampaignCounterRecompute:
    """Recomputation stays source-of-truth, but must not scan rows in Python."""

    def test_adjustments_are_summed_in_the_database(self) -> None:
        """The old python loop fetched every adjustment row on every payment event."""
        from apps.madadkar.choices import FinancialAdjustmentStatus, FinancialAdjustmentType
        from apps.madadkar.models import CampaignFinancialAdjustment
        from apps.madadkar.services import _sync_campaign_counters
        from tests.factories.madadkar import PublishedCampaignFactory

        campaign = PublishedCampaignFactory()
        CampaignFinancialAdjustment.objects.bulk_create([
            CampaignFinancialAdjustment(
                campaign=campaign, amount=100,
                adjustment_type=FinancialAdjustmentType.CREDIT,
                status=FinancialAdjustmentStatus.APPLIED, reason=f"r{index}",
            )
            for index in range(25)
        ])

        with CaptureQueriesContext(connection) as ctx:
            _sync_campaign_counters(campaign=campaign)

        adjustment_reads = [
            query["sql"] for query in ctx.captured_queries
            if "financialadjustment" in query["sql"].lower()
        ]
        assert len(adjustment_reads) == 1
        assert "SUM" in adjustment_reads[0].upper()

    def test_participation_counters_share_one_scan(self) -> None:
        """Reserved shares, paid amount and unique users come from one aggregate."""
        from apps.madadkar.services import _sync_campaign_counters
        from tests.factories.madadkar import PaidParticipationFactory, PublishedCampaignFactory

        campaign = PublishedCampaignFactory()
        PaidParticipationFactory(campaign=campaign)

        with CaptureQueriesContext(connection) as ctx:
            _sync_campaign_counters(campaign=campaign)

        # فقط کوئری‌هایی که *از* جدول مشارکت می‌خوانند؛ کوئری refund هم به آن
        # JOIN می‌زند ولی FROM آن جدول دیگری است.
        participation_reads = [
            query["sql"] for query in ctx.captured_queries
            if 'FROM "madadkar_participation"' in query["sql"]
        ]
        assert len(participation_reads) == 1
        assert participation_reads[0].count("FILTER (WHERE") == 3

    def test_credit_and_debit_signs_are_preserved(self) -> None:
        """The DB-side Case/When must match the python signed_amount property."""
        from apps.madadkar.choices import FinancialAdjustmentStatus, FinancialAdjustmentType
        from apps.madadkar.models import CampaignFinancialAdjustment
        from apps.madadkar.services import _sync_campaign_counters
        from tests.factories.madadkar import PaidParticipationFactory, PublishedCampaignFactory

        campaign = PublishedCampaignFactory()
        participation = PaidParticipationFactory(campaign=campaign)
        baseline = participation.total_amount

        CampaignFinancialAdjustment.objects.create(
            campaign=campaign, amount=5_000,
            adjustment_type=FinancialAdjustmentType.CREDIT,
            status=FinancialAdjustmentStatus.APPLIED, reason="credit",
        )
        CampaignFinancialAdjustment.objects.create(
            campaign=campaign, amount=2_000,
            adjustment_type=FinancialAdjustmentType.DEBIT,
            status=FinancialAdjustmentStatus.APPLIED, reason="debit",
        )
        _sync_campaign_counters(campaign=campaign)
        campaign.refresh_from_db()
        assert campaign.purchased_amount == baseline + 5_000 - 2_000

    def test_unapplied_adjustments_are_ignored(self) -> None:
        """Only APPLIED adjustments may move the money counter."""
        from apps.madadkar.choices import FinancialAdjustmentStatus, FinancialAdjustmentType
        from apps.madadkar.models import CampaignFinancialAdjustment
        from apps.madadkar.services import _sync_campaign_counters
        from tests.factories.madadkar import PaidParticipationFactory, PublishedCampaignFactory

        campaign = PublishedCampaignFactory()
        participation = PaidParticipationFactory(campaign=campaign)
        CampaignFinancialAdjustment.objects.create(
            campaign=campaign, amount=9_999,
            adjustment_type=FinancialAdjustmentType.CREDIT,
            status=FinancialAdjustmentStatus.PENDING_REVIEW, reason="pending",
        )
        _sync_campaign_counters(campaign=campaign)
        campaign.refresh_from_db()
        assert campaign.purchased_amount == participation.total_amount
